from __future__ import annotations

import argparse
from pathlib import Path

from harness_lib import write_json


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect workbook worksheets for eligibility metadata.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--workbook-file-id", default="WORKBOOK")
    parser.add_argument("--project-id")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    if args.workbook.suffix.lower() == ".xls":
        write_json(args.output, {
            "adjudication_id": "workbook-sheet-adjudication",
            "status": "UNSUPPORTED_XLS_PARSER_REQUIRED",
            "worksheets": []
        })
        return

    from openpyxl import load_workbook

    wb = load_workbook(args.workbook, read_only=False, data_only=False)
    external_links = [str(x) for x in getattr(wb, "_external_links", [])]
    rows = []
    for idx, ws in enumerate(wb.worksheets, start=1):
        values = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            values.extend("" if v is None else str(v) for v in row[:10])
        joined = " ".join(values)
        stale = "CURRENT_OR_UNVERIFIED"
        eligibility = "HUMAN_REVIEW_REQUIRED"
        if args.project_id and args.project_id not in joined:
            stale = "INSUFFICIENT_IDENTITY"
        if ws.sheet_state != "visible":
            eligibility = "HUMAN_REVIEW_REQUIRED"
        rows.append({
            "sheet_id": f"{args.workbook_file_id}-SHEET-{idx:03d}",
            "workbook_file_id": args.workbook_file_id,
            "sheet_name": ws.title,
            "visibility": ws.sheet_state,
            "used_range": ws.calculate_dimension(),
            "print_area": str(ws.print_area) if ws.print_area else None,
            "named_ranges": [],
            "external_links": external_links,
            "detected_project_id": args.project_id if args.project_id and args.project_id in joined else None,
            "detected_customer": None,
            "detected_project_name": None,
            "detected_date": None,
            "detected_revision": None,
            "stale_template_status": stale,
            "evidence": joined[:500],
            "generator_eligibility": eligibility,
            "approved_by": None,
            "approval_time": None
        })
    write_json(args.output, {
        "adjudication_id": "workbook-sheet-adjudication",
        "created_at": "LOCAL",
        "status": "INSPECTED",
        "worksheets": rows
    })


if __name__ == "__main__":
    main()

