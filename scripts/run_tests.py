from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path

from harness_lib import classify_path, read_json, repo_root, sha256_file, sha256_json, validate, validate_file
from harness_lib import write_json


ROOT = repo_root()
PY = sys.executable


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def run(cmd: list[str]) -> None:
    print("$", " ".join(cmd))
    subprocess.run(cmd, cwd=ROOT, check=True)


def test_json_schemas_parse() -> None:
    for path in (ROOT / "schemas").glob("*.schema.json"):
        read_json(path)


def test_declared_json_artifacts_validate() -> None:
    pairs = [
        ("manifests/source_manifest_initial.json", "schemas/source_manifest.schema.json"),
        ("manifests/workbook_sheet_adjudication_initial.json", "schemas/workbook_sheet_adjudication.schema.json"),
        ("manifests/generator_input_manifest_empty.json", "schemas/generator_input_manifest.schema.json"),
        ("rules/source_authority_profile.json", "schemas/source_authority_profile.schema.json"),
        ("rules/derivation_rules/bootstrap_dimension_rule.json", "schemas/derivation_rule.schema.json"),
        ("rules/constraint_rules/bootstrap_cross_pdf_rule.json", "schemas/constraint_rule.schema.json"),
        ("evals/grading_profiles/plc_layout_v1.json", "schemas/grading_result.schema.json"),
    ]
    # The grading profile has its own structure; check weights separately instead of grading_result schema.
    for instance, schema in pairs[:-1]:
        errors = validate_file(ROOT / instance, ROOT / schema)
        assert_true(not errors, f"{instance} failed {schema}: {errors}")


def test_utf8_fixture_labels() -> None:
    text = (ROOT / "evals/fixtures/utf8_labels.txt").read_text(encoding="utf-8")
    for label in ["合約", "材料清單", "生管文件", "生管課用圖", "生管用圖", "沖孔", "鈑金", "電機施工圖"]:
        assert_true(label in text, f"missing label {label}")


def test_forbidden_classification() -> None:
    forbidden = classify_path("案號/電機施工圖/01_生管課用圖_ABC.pdf")
    assert_true(forbidden["generator_eligibility"] == "FORBIDDEN", "target-like electrical drawing must be forbidden")
    allowedish = classify_path("案號/合約_ABC.xlsx")
    assert_true(allowedish["primary_role"] == "allowed_contract_workbook", "contract workbook label should classify")


def test_grading_weights() -> None:
    profile = read_json(ROOT / "evals/grading_profiles/plc_layout_v1.json")
    assert_true(sum(profile["weights"].values()) == 100, "grading weights must sum to 100")
    assert_true(profile["highest_automated_status"] == "READY_FOR_PRIVATE_PREVIEW", "highest automated status guard changed")


def test_evaluator_sensitivity_monotonicity() -> None:
    from evaluator_scoring import compute_score

    base = {
        "run_id": "SENS-BASE",
        "validity_inputs": {"required_outputs_present": True, "hard_gate_failures": []},
        "coverage": {"scorable_elements": 38, "total_elements": 100},
        "dimension_items": {
            "source_selection_provenance_conflict": [
                {"item_id": "source_bundle", "score": 14, "max_score": 20, "classification": "EXPLICIT_IN_ALLOWED_INPUT"}
            ],
            "panel_schedule_enclosure_facts": [
                {"item_id": "panel_facts", "score": 5, "max_score": 15, "classification": "UNAVAILABLE_IN_ALLOWED_INPUT"}
            ],
            "device_bom_quantity_tag_fidelity": [
                {"item_id": "device_fidelity", "score": 8, "max_score": 20, "classification": "UNAVAILABLE_IN_ALLOWED_INPUT"}
            ],
            "shared_geometry_cross_pdf_consistency": [
                {"item_id": "cross_pdf", "score": 8, "max_score": 15, "classification": "DESIGN_CHOICE_WITH_CONSTRAINTS"}
            ],
            "production_drawing_quality": [
                {"item_id": "production", "score": 3, "max_score": 10, "classification": "DESIGN_CHOICE_WITH_CONSTRAINTS"}
            ],
            "punch_drawing_quality": [
                {"item_id": "punch", "score": 2, "max_score": 10, "classification": "DESIGN_CHOICE_WITH_CONSTRAINTS"}
            ],
            "sheetmetal_drawing_quality": [
                {"item_id": "sheetmetal", "score": 2, "max_score": 10, "classification": "DESIGN_CHOICE_WITH_CONSTRAINTS"}
            ],
        },
        "findings": [
            {"finding_id": "F1", "severity": "HIGH", "dedupe_key": "missing_exact_dimensions", "classification": "UNAVAILABLE_IN_ALLOWED_INPUT"},
            {"finding_id": "F2", "severity": "HIGH", "dedupe_key": "reference_layout_unavailable", "classification": "UNAVAILABLE_IN_ALLOWED_INPUT"},
            {"finding_id": "F3", "severity": "HIGH", "dedupe_key": "schematic_minimal", "classification": "DESIGN_CHOICE_WITH_CONSTRAINTS"},
        ],
    }

    baseline = compute_score(base)
    assert_true(baseline["quality_score"] == 42, "baseline score should be arithmetic sum of dimension items")
    assert_true(baseline["scorable_coverage"] == 38, "coverage should come from the explicit denominator")

    missing = json.loads(json.dumps(base))
    missing["validity_inputs"]["required_outputs_present"] = False
    assert_true(compute_score(missing)["validity"] == "FAIL", "missing output must fail validity")

    corrected = json.loads(json.dumps(base))
    corrected["dimension_items"]["panel_schedule_enclosure_facts"][0]["score"] = 9
    assert_true(compute_score(corrected)["quality_score"] > baseline["quality_score"], "proven correction must improve score")

    worse = json.loads(json.dumps(base))
    worse["dimension_items"]["device_bom_quantity_tag_fidelity"][0]["score"] = 4
    assert_true(compute_score(worse)["quality_score"] < baseline["quality_score"], "new high defect must worsen score")

    unsupported = json.loads(json.dumps(base))
    unsupported["validity_inputs"]["hard_gate_failures"] = ["INVENTED_CRITICAL_VALUE"]
    gated = compute_score(unsupported)
    assert_true(gated["validity"] == "FAIL" and gated["quality_score"] == 0, "hard gate must override numerical score")

    no_scorable = json.loads(json.dumps(base))
    no_scorable["coverage"] = {"scorable_elements": 0, "total_elements": 0}
    low = compute_score(no_scorable)
    assert_true(low["evidence_strength"] == "INSUFFICIENT_COVERAGE", "zero scorable elements must not look like an ordinary score")

    duplicate = json.loads(json.dumps(base))
    duplicate["findings"].append(dict(duplicate["findings"][0], finding_id="F1_DUP"))
    deduped = compute_score(duplicate)
    assert_true(deduped["high_findings"] == baseline["high_findings"], "duplicate findings must not add accidental penalties")

    same_total = json.loads(json.dumps(base))
    same_total["dimension_items"]["production_drawing_quality"][0]["score"] = 4
    same_total["dimension_items"]["punch_drawing_quality"][0]["score"] = 1
    alt = compute_score(same_total)
    assert_true(alt["quality_score"] == baseline["quality_score"], "fixture should preserve same total")
    assert_true(alt["dimension_scores"] != baseline["dimension_scores"], "same total must preserve different dimension vectors")


def test_synthetic_render_grade() -> None:
    work_dir = ROOT / "tmp" / "phase0_test_harness"
    run([PY, "scripts/eval_harness.py", "--work-dir", str(work_dir)])
    grading = read_json(work_dir / "grading_result.json")
    errors = validate(grading, read_json(ROOT / "schemas/grading_result.schema.json"))
    assert_true(not errors, f"grading_result schema errors: {errors}")
    assert_true(grading["validity"] == "PASS", "synthetic PDF package should pass")


def test_positive_bundle_and_contamination_scan() -> None:
    work_dir = ROOT / "tmp" / "bundle_test"
    work_dir.mkdir(parents=True, exist_ok=True)
    source_manifest = work_dir / "source_manifest.json"
    sheet_adjudication = work_dir / "sheet_adjudication.json"
    bundle_dir = work_dir / "bundle"
    generator_manifest = work_dir / "generator_input_manifest.json"
    contamination = work_dir / "contamination.json"
    write_json(source_manifest, {
        "manifest_id": "bundle-test-source",
        "files": [
            {
                "file_id": "FILE-ALLOWED",
                "relative_path": "1150101/合約_1150101.xlsx",
                "file_name": "合約_1150101.xlsx",
                "primary_role": "allowed_contract_workbook",
                "generator_eligibility": "ALLOWED"
            },
            {
                "file_id": "FILE-FORBIDDEN",
                "relative_path": "1150101/電機施工圖/target.pdf",
                "file_name": "target.pdf",
                "primary_role": "forbidden_electrical_drawing",
                "generator_eligibility": "FORBIDDEN"
            }
        ],
        "worksheets": []
    })
    write_json(sheet_adjudication, {
        "adjudication_id": "bundle-test-sheets",
        "worksheets": [
            {"sheet_id": "SHEET-ALLOWED", "generator_eligibility": "ALLOWED", "stale_template_status": "CURRENT_PROJECT_ID_MATCH"},
            {"sheet_id": "SHEET-STALE", "generator_eligibility": "FORBIDDEN", "stale_template_status": "STALE_TEMPLATE_SHEET"}
        ]
    })
    run([PY, "scripts/build_generator_bundle.py", "--source-manifest", str(source_manifest), "--sheet-adjudication", str(sheet_adjudication), "--bundle-dir", str(bundle_dir), "--output-manifest", str(generator_manifest), "--run-id", "BUNDLE-TEST", "--project-id", "1150101"])
    manifest = read_json(generator_manifest)
    assert_true(manifest["allowed_files"] == ["FILE-ALLOWED"], "only positive allowed file should enter manifest")
    assert_true(manifest["allowed_sheets"] == ["SHEET-ALLOWED"], "stale sheet must be excluded")
    run([PY, "scripts/scan_generator_contamination.py", "--bundle-dir", str(bundle_dir), "--manifest", str(generator_manifest), "--output", str(contamination)])
    scan = read_json(contamination)
    assert_true(scan["status"] == "PASS", "clean positive bundle should pass contamination scan")


def create_source_guard_fixture(root: Path):
    from openpyxl import Workbook

    root.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "核准輸入"
    ws["A1"] = "工程編號"
    ws["B1"] = "1159999"
    ws["A2"] = "客戶"
    ws["B2"] = "目前客戶"
    ws["A3"] = "公式"
    ws["B3"] = "=DeniedSheet!A1"
    denied = wb.create_sheet("DeniedSheet")
    denied["A1"] = "denied"
    hidden = wb.create_sheet("hidden舊資料")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "1140001 舊客戶"
    sentinel = wb.create_sheet("veryHidden完成圖")
    sentinel.sheet_state = "veryHidden"
    sentinel["A1"] = "生管課用圖 completed target sentinel"
    try:
        from openpyxl.workbook.defined_name import DefinedName

        wb.defined_names.add(DefinedName("DeniedRange", attr_text="'DeniedSheet'!$A$1"))
    except Exception:
        pass
    path = root / "合約_1159999.xlsx"
    wb.save(path)
    macro_path = root / "合約_1159999.xlsm"
    wb.save(macro_path)
    legacy_path = root / "舊格式_1159999.xls"
    legacy_path.write_bytes(b"legacy xls placeholder")
    return path, macro_path, legacy_path


def test_source_guard_fail_closed_decisions() -> None:
    import source_guard

    fixture_root = ROOT / "tmp" / "source_guard_fixtures" / "SRC"
    workbook, macro_workbook, legacy = create_source_guard_fixture(fixture_root)
    old_root = source_guard.SOURCE_ROOT_PATH
    source_guard.SOURCE_ROOT_PATH = fixture_root
    try:
        file_row = {
            "file_id": "FILE-1",
            "source_root": "SRC-ALL-PROJECTS",
            "project_id": "1159999",
            "project_name": "1159999測試-目前客戶",
            "customer": "目前客戶",
            "relative_path": workbook.name,
            "file_name": workbook.name,
            "extension": ".xlsx",
            "size_bytes": str(workbook.stat().st_size),
            "sha256": __import__("hashlib").sha256(workbook.read_bytes()).hexdigest().upper(),
            "primary_role": "allowed_contract_workbook",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "parse_status": "PARSED",
            "absolute_path": str(workbook),
            "duplicate_of": "",
        }
        visible_sheet = {
            "sheet_id": "SHEET-1",
            "sheet_name": "核准輸入",
            "visibility": "visible",
            "stale_template_status": "CURRENT_PROJECT_ID_MATCH",
            "stale_evidence": "1159999 目前客戶",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "external_links": "external.xlsx",
            "named_ranges": "DeniedRange",
            "parse_status": "PARSED",
        }
        hidden_sheet = dict(visible_sheet, sheet_id="SHEET-2", sheet_name="hidden舊資料", visibility="hidden", stale_template_status="INSUFFICIENT_IDENTITY", stale_evidence="1140001 舊客戶")
        sentinel_sheet = dict(visible_sheet, sheet_id="SHEET-3", sheet_name="veryHidden完成圖", visibility="veryHidden", stale_template_status="FORBIDDEN_SOURCE", stale_evidence="生管課用圖")
        visible_decision = source_guard.decide_source_item(file_row, visible_sheet)
        hidden_decision = source_guard.decide_source_item(file_row, hidden_sheet)
        sentinel_decision = source_guard.decide_source_item(file_row, sentinel_sheet)
        assert_true(visible_decision["proposed_decision"] == "QUARANTINED", "formula/external/named-range dependencies must quarantine")
        assert_true("FORMULA_REFERENCES_OTHER_SHEET" in visible_decision["reason_codes"], "formula dependency reason missing")
        assert_true(hidden_decision["proposed_decision"] == "QUARANTINED", "hidden sheet must not be approved")
        assert_true(sentinel_decision["proposed_decision"] == "AUTO_DENIED", "completed-reference sentinel must auto-deny")
        assert_true(visible_decision["worksheet_fingerprint"], "visible supported sheet must have deterministic fingerprint")
        again = source_guard.decide_source_item(file_row, visible_sheet)
        assert_true(again["worksheet_fingerprint"] == visible_decision["worksheet_fingerprint"], "worksheet fingerprint must be deterministic")

        macro_row = dict(file_row, file_id="FILE-MACRO", relative_path=macro_workbook.name, file_name=macro_workbook.name, extension=".xlsm", size_bytes=str(macro_workbook.stat().st_size), sha256=__import__("hashlib").sha256(macro_workbook.read_bytes()).hexdigest().upper(), absolute_path=str(macro_workbook))
        macro_decision = source_guard.decide_source_item(macro_row, dict(visible_sheet, external_links="", named_ranges=""))
        assert_true(macro_decision["proposed_decision"] in {"QUARANTINED", "AUTO_DENIED"}, "macro-enabled workbook must fail closed")

        legacy_row = dict(file_row, file_id="FILE-XLS", relative_path=legacy.name, file_name=legacy.name, extension=".xls", size_bytes=str(legacy.stat().st_size), sha256=__import__("hashlib").sha256(legacy.read_bytes()).hexdigest().upper(), absolute_path=str(legacy))
        legacy_decision = source_guard.decide_source_item(legacy_row, None)
        assert_true(legacy_decision["proposed_decision"] == "PARSER_REQUIRED", "legacy xls must be parser-required")

        tampered = dict(file_row, sha256="0" * 64)
        tampered_decision = source_guard.decide_source_item(tampered, visible_sheet)
        assert_true(tampered_decision["proposed_decision"] == "AUTO_DENIED", "source hash mutation must deny")
    finally:
        source_guard.SOURCE_ROOT_PATH = old_root


def test_source_approval_and_bundle_fail_closed() -> None:
    fixture_root = ROOT / "tmp" / "source_guard_approval"
    workbook, _, _ = create_source_guard_fixture(fixture_root)
    file_sha = __import__("hashlib").sha256(workbook.read_bytes()).hexdigest().upper()
    import source_guard

    old_root = source_guard.SOURCE_ROOT_PATH
    source_guard.SOURCE_ROOT_PATH = fixture_root
    try:
        file_row = {
            "file_id": "FILE-APPROVE",
            "source_root": "SRC-ALL-PROJECTS",
            "project_id": "1159999",
            "project_name": "1159999測試-目前客戶",
            "customer": "目前客戶",
            "relative_path": workbook.name,
            "file_name": workbook.name,
            "extension": ".xlsx",
            "size_bytes": str(workbook.stat().st_size),
            "sha256": file_sha,
            "primary_role": "allowed_contract_workbook",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "parse_status": "PARSED",
            "absolute_path": str(workbook),
            "duplicate_of": "",
        }
        sheet_row = {
            "sheet_id": "SHEET-APPROVE",
            "sheet_name": "核准輸入",
            "visibility": "visible",
            "stale_template_status": "CURRENT_PROJECT_ID_MATCH",
            "stale_evidence": "1159999 目前客戶",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "external_links": "",
            "named_ranges": "",
            "parse_status": "PARSED",
        }
        decision = source_guard.decide_source_item(file_row, sheet_row)
    finally:
        source_guard.SOURCE_ROOT_PATH = old_root
    decision["proposed_decision"] = "CANDIDATE"
    decisions_csv = fixture_root / "decisions.csv"
    fields = ["decision_id", "project_id", "file_id", "sheet_id", "file_sha256", "worksheet_fingerprint", "proposed_decision"]
    import csv

    with decisions_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(decision.keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerow({k: "|".join(v) if isinstance(v, list) else v for k, v in decision.items()})
    blank_decisions = fixture_root / "blank_human_decisions.csv"
    with blank_decisions.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["decision_id", "project_id", "file_id", "sheet_id", "file_sha256", "worksheet_fingerprint", "human_decision", "reviewer", "notes"])
        writer.writeheader()
        writer.writerow({"decision_id": decision["decision_id"], "project_id": "1159999", "file_id": "FILE-APPROVE", "sheet_id": "SHEET-APPROVE", "file_sha256": file_sha, "worksheet_fingerprint": decision["worksheet_fingerprint"], "human_decision": "", "reviewer": "", "notes": ""})
    failed = subprocess.run([PY, "scripts/validate_source_approval.py", "--decisions", str(decisions_csv), "--human-decisions", str(blank_decisions), "--output", str(fixture_root / "approval_fail.json")], cwd=ROOT)
    assert_true(failed.returncode != 0, "blank human decision must fail")

    approve_csv = fixture_root / "approve.csv"
    with approve_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["decision_id", "project_id", "file_id", "sheet_id", "file_sha256", "worksheet_fingerprint", "human_decision", "reviewer", "notes"])
        writer.writeheader()
        writer.writerow({"decision_id": decision["decision_id"], "project_id": "1159999", "file_id": "FILE-APPROVE", "sheet_id": "SHEET-APPROVE", "file_sha256": file_sha, "worksheet_fingerprint": "bad", "human_decision": "HUMAN_APPROVED", "reviewer": "tester", "notes": ""})
    tamper = subprocess.run([PY, "scripts/validate_source_approval.py", "--decisions", str(decisions_csv), "--human-decisions", str(approve_csv), "--output", str(fixture_root / "approval_tamper.json")], cwd=ROOT)
    assert_true(tamper.returncode != 0, "tampered worksheet fingerprint must fail")


def synthetic_signed_authority_decision(selected_choice_ids: list[str] | None = None) -> dict:
    packet_path = ROOT / "reports" / "sheetmetal-v1" / "source-rule-approval" / "smv1_source_rule_authority_decision_packet.json"
    template_path = ROOT / "reports" / "sheetmetal-v1" / "source-rule-approval" / "smv1_signed_authority_decision_template.json"
    return {
        "schema_version": "sheetmetal-v1.signed_authority_decision.v1",
        "decision_id": "TEST-SIGNED-AUTHORITY",
        "active_goal": "SHEETMETAL_FIRST_MODULAR_PANEL_MODEL_V1",
        "bound_decision_packet": {
            "json_path": "reports/sheetmetal-v1/source-rule-approval/smv1_source_rule_authority_decision_packet.json",
            "json_sha256": sha256_file(packet_path),
            "template_path": "reports/sheetmetal-v1/source-rule-approval/smv1_signed_authority_decision_template.json",
            "template_sha256": sha256_file(template_path),
        },
        "selected_choice_ids": selected_choice_ids or ["A", "C"],
        "signed_by": "test authority",
        "signed_at": "2026-06-24",
        "signed_statement": "I authorize the following SMV1 source/rule authority choices: A, C.",
        "non_negotiable_constraints_acknowledged": [
            "no_production_approval",
            "no_customer_pdf_dxf_dwg_generation_before_authorized_renderer_gate",
            "no_completed_reference_inference",
            "no_post_design_label_generator_authority",
            "no_private_model_identifier_public_lookup",
            "no_source_root_mutation",
            "no_private_artifact_staging",
            "tests_before_any_fix",
            "independent_audit_before_model_promotion",
        ],
        "implementation_requires_regression_tests": True,
        "independent_audit_before_model_promotion": True,
        "customer_pdf_dxf_dwg_generation_authorized": False,
        "production_approval_declared": False,
    }


def test_signed_authority_decision_validator_fail_closed() -> None:
    work = ROOT / "tmp" / "signed_authority_decision_validator"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    base_decision = synthetic_signed_authority_decision()
    valid_path = work / "valid_decision.json"
    valid_out = work / "valid_result.json"
    write_json(valid_path, base_decision)
    passed = subprocess.run([PY, "scripts/validate_signed_authority_decision.py", "--decision", str(valid_path), "--output", str(valid_out)], cwd=ROOT, capture_output=True, text=True)
    assert_true(passed.returncode == 0, f"valid signed authority decision should pass: {passed.stdout} {passed.stderr}")
    passed_data = read_json(valid_out)
    assert_true(passed_data["status"] == "PASS", "valid signed authority decision output must pass")
    assert_true(passed_data["selected_choice_ids"] == ["A", "C"], "accepted choice IDs must be preserved")
    assert_true(passed_data["implementation_can_start"] is False, "signed authority alone must not start implementation")

    reject_all_conflict = json.loads(json.dumps(base_decision))
    reject_all_conflict["selected_choice_ids"] = ["A", "D"]
    reject_all_path = work / "reject_all_conflict.json"
    write_json(reject_all_path, reject_all_conflict)
    failed = subprocess.run([PY, "scripts/validate_signed_authority_decision.py", "--decision", str(reject_all_path), "--output", str(work / "reject_all_result.json")], cwd=ROOT, capture_output=True, text=True)
    assert_true(failed.returncode != 0, "reject-all must be mutually exclusive")
    assert_true("REJECT_ALL_MUTUALLY_EXCLUSIVE" in failed.stdout + failed.stderr, "reject-all conflict reason missing")

    unsafe_flags = json.loads(json.dumps(base_decision))
    unsafe_flags["customer_pdf_dxf_dwg_generation_authorized"] = True
    unsafe_flags["production_approval_declared"] = True
    unsafe_path = work / "unsafe_flags.json"
    write_json(unsafe_path, unsafe_flags)
    failed = subprocess.run([PY, "scripts/validate_signed_authority_decision.py", "--decision", str(unsafe_path), "--output", str(work / "unsafe_result.json")], cwd=ROOT, capture_output=True, text=True)
    assert_true(failed.returncode != 0, "drawing generation and production approval flags must fail")
    assert_true("CUSTOMER_DRAWING_GENERATION_NOT_ALLOWED_AT_DECISION_GATE" in failed.stdout + failed.stderr, "drawing-generation guard reason missing")
    assert_true("PRODUCTION_APPROVAL_FORBIDDEN" in failed.stdout + failed.stderr, "production-approval guard reason missing")

    hash_mismatch = json.loads(json.dumps(base_decision))
    hash_mismatch["bound_decision_packet"]["json_sha256"] = "0" * 64
    hash_path = work / "hash_mismatch.json"
    write_json(hash_path, hash_mismatch)
    failed = subprocess.run([PY, "scripts/validate_signed_authority_decision.py", "--decision", str(hash_path), "--output", str(work / "hash_result.json")], cwd=ROOT, capture_output=True, text=True)
    assert_true(failed.returncode != 0, "bound decision packet hash mismatch must fail")
    assert_true("BOUND_DECISION_PACKET_HASH_MISMATCH" in failed.stdout + failed.stderr, "hash mismatch reason missing")


def test_signed_authority_decision_intake_routing() -> None:
    work = ROOT / "tmp" / "signed_authority_decision_intake"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    accepted_decision = synthetic_signed_authority_decision(["A", "B"])
    accepted_path = work / "accepted_decision.json"
    accepted_out = work / "accepted_intake.json"
    write_json(accepted_path, accepted_decision)
    passed = subprocess.run([PY, "scripts/prepare_signed_authority_intake.py", "--decision", str(accepted_path), "--output", str(accepted_out)], cwd=ROOT, capture_output=True, text=True)
    assert_true(passed.returncode == 0, f"accepted-lane intake should pass: {passed.stdout} {passed.stderr}")
    accepted_data = read_json(accepted_out)
    assert_true(accepted_data["status"] == "PASS", "accepted-lane intake status must pass")
    assert_true(accepted_data["selected_choice_ids"] == ["A", "B"], "accepted choices must be normalized")
    assert_true(accepted_data["next_action"] == "ADD_REGRESSION_TESTS_BEFORE_ACCEPTED_AUTHORITY_LANE_FIX", "accepted lanes must route to test-before-fix gate")
    assert_true(accepted_data["implementation_authorized"] is False, "intake must not authorize implementation directly")
    assert_true(accepted_data["customer_pdf_dxf_dwg_generation_authorized"] is False, "intake must not authorize customer drawings")
    assert_true(accepted_data["production_approval_declared"] is False, "intake must not declare production approval")

    rejected_decision = synthetic_signed_authority_decision(["D"])
    rejected_path = work / "reject_all_decision.json"
    rejected_out = work / "reject_all_intake.json"
    write_json(rejected_path, rejected_decision)
    passed = subprocess.run([PY, "scripts/prepare_signed_authority_intake.py", "--decision", str(rejected_path), "--output", str(rejected_out)], cwd=ROOT, capture_output=True, text=True)
    assert_true(passed.returncode == 0, f"reject-all intake should pass: {passed.stdout} {passed.stderr}")
    rejected_data = read_json(rejected_out)
    assert_true(rejected_data["reject_all_authority_lanes"] is True, "reject-all intake must preserve terminal branch")
    assert_true(rejected_data["next_action"] == "ENTER_TERMINAL_CANDIDATE_REVIEW", "reject-all must route to terminal-candidate review")

    invalid_decision = synthetic_signed_authority_decision(["A", "D"])
    invalid_path = work / "invalid_decision.json"
    invalid_out = work / "invalid_intake.json"
    write_json(invalid_path, invalid_decision)
    failed = subprocess.run([PY, "scripts/prepare_signed_authority_intake.py", "--decision", str(invalid_path), "--output", str(invalid_out)], cwd=ROOT, capture_output=True, text=True)
    assert_true(failed.returncode != 0, "invalid signed decision intake must fail closed")
    invalid_data = read_json(invalid_out)
    assert_true(invalid_data["status"] == "FAIL", "invalid intake output must fail")
    assert_true(invalid_data["next_action"] == "WAIT_FOR_SIGNED_HUMAN_SOURCE_RULE_AUTHORITY_DECISION", "invalid intake must keep waiting for a valid signed decision")


def test_signed_authority_decision_draft_scaffold_fail_closed() -> None:
    work = ROOT / "tmp" / "signed_authority_decision_draft"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    draft_path = work / "draft_decision.json"
    result = subprocess.run([PY, "scripts/prepare_signed_authority_decision_draft.py", "--output", str(draft_path)], cwd=ROOT, capture_output=True, text=True)
    assert_true(result.returncode == 0, f"unsigned decision draft should be generated: {result.stdout} {result.stderr}")
    draft = read_json(draft_path)
    assert_true(draft["schema_version"] == "sheetmetal-v1.signed_authority_decision.v1", "draft must use signed decision schema version")
    assert_true(draft["selected_choice_ids"] == [], "draft must not select authority choices")
    assert_true(draft["signed_by"] == "", "draft must not invent signer")
    assert_true(draft["signed_at"] == "YYYY-MM-DD", "draft must require signer date replacement")
    assert_true(draft["customer_pdf_dxf_dwg_generation_authorized"] is False, "draft must not authorize customer drawings")
    assert_true(draft["production_approval_declared"] is False, "draft must not declare production approval")

    packet_path = ROOT / draft["bound_decision_packet"]["json_path"]
    template_path = ROOT / draft["bound_decision_packet"]["template_path"]
    assert_true(draft["bound_decision_packet"]["json_sha256"] == sha256_file(packet_path), "draft must bind current decision packet hash")
    assert_true(draft["bound_decision_packet"]["template_sha256"] == sha256_file(template_path), "draft must bind current signed-decision template hash")

    validation_out = work / "draft_validation.json"
    failed = subprocess.run([PY, "scripts/validate_signed_authority_decision.py", "--decision", str(draft_path), "--output", str(validation_out)], cwd=ROOT, capture_output=True, text=True)
    assert_true(failed.returncode != 0, "unsigned draft must fail signed decision validation")
    validation = read_json(validation_out)
    assert_true(validation["status"] == "FAIL", "unsigned draft validation status must fail")
    assert_true("NO_AUTHORITY_CHOICES_SELECTED" in validation["errors"], "unsigned draft must not pass without selected choices")
    assert_true("MISSING_SIGNER" in validation["errors"], "unsigned draft must not pass without signer")
    assert_true("INVALID_SIGNED_AT_DATE" in validation["errors"], "unsigned draft must require a real signed date")
    assert_true(all(check["status"] == "PASS" for check in validation["hash_checks"]), "draft hash bindings must be valid")


def test_signed_authority_decision_submission_package() -> None:
    work = ROOT / "tmp" / "signed_authority_decision_submission"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    accepted_decision = synthetic_signed_authority_decision(["A", "C"])
    accepted_path = work / "accepted_decision.json"
    accepted_out = work / "accepted_submission"
    write_json(accepted_path, accepted_decision)
    passed = subprocess.run(
        [
            PY,
            "scripts/process_signed_authority_decision.py",
            "--decision",
            str(accepted_path),
            "--output-dir",
            str(accepted_out),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )
    assert_true(passed.returncode == 0, f"valid signed decision package should pass: {passed.stdout} {passed.stderr}")
    validation = read_json(accepted_out / "validation.json")
    intake = read_json(accepted_out / "intake.json")
    summary = read_json(accepted_out / "submission_summary.json")
    assert_true(validation["status"] == "PASS", "submission package validation must pass for valid decision")
    assert_true(intake["status"] == "PASS", "submission package intake must pass for valid decision")
    assert_true(summary["status"] == "SIGNED_AUTHORITY_DECISION_VALIDATED_INTAKE_READY", "accepted-lane submission status must be intake-ready")
    assert_true(summary["selected_choice_ids"] == ["A", "C"], "submission package must preserve accepted choices")
    assert_true(summary["next_action"] == "ADD_REGRESSION_TESTS_BEFORE_ACCEPTED_AUTHORITY_LANE_FIX", "accepted submission must route to test-before-fix gate")
    assert_true(summary["implementation_authorized"] is False, "submission package must not directly authorize implementation")
    assert_true(summary["customer_pdf_dxf_dwg_generation_authorized"] is False, "submission package must not authorize customer drawings")
    assert_true(summary["production_approval_declared"] is False, "submission package must not declare production approval")
    assert_true(all(check["status"] == "PASS" for check in summary["hash_checks"]), "submission package must preserve passing hash checks")

    draft_path = work / "unsigned_draft.json"
    draft_result = subprocess.run([PY, "scripts/prepare_signed_authority_decision_draft.py", "--output", str(draft_path)], cwd=ROOT, capture_output=True, text=True)
    assert_true(draft_result.returncode == 0, f"unsigned draft setup should pass: {draft_result.stdout} {draft_result.stderr}")
    draft_out = work / "draft_submission"
    failed = subprocess.run(
        [
            PY,
            "scripts/process_signed_authority_decision.py",
            "--decision",
            str(draft_path),
            "--output-dir",
            str(draft_out),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )
    assert_true(failed.returncode != 0, "unsigned draft submission package must fail closed")
    draft_summary = read_json(draft_out / "submission_summary.json")
    assert_true(draft_summary["status"] == "SIGNED_AUTHORITY_DECISION_INVALID_FAIL_CLOSED", "invalid submission summary must fail closed")
    assert_true(draft_summary["next_action"] == "WAIT_FOR_SIGNED_HUMAN_SOURCE_RULE_AUTHORITY_DECISION", "invalid submission must keep waiting")
    assert_true("NO_AUTHORITY_CHOICES_SELECTED" in draft_summary["validation_errors"], "invalid submission must surface missing choice")
    assert_true("MISSING_SIGNER" in draft_summary["validation_errors"], "invalid submission must surface missing signer")
    assert_true("INVALID_SIGNED_AT_DATE" in draft_summary["validation_errors"], "invalid submission must surface missing date")


def test_bundle_verifier_rejects_leaks_and_traversal() -> None:
    bundle = ROOT / "tmp" / "bad_bundle"
    bundle.mkdir(parents=True, exist_ok=True)
    write_json(bundle / "bundle_manifest.json", {"bundle_id": "bad", "project_id": "P", "policy_version": "source_guard_policy_v1", "status": "BUILT", "artifacts": [{"artifact_id": "A", "path": "sanitized_inputs/01_生管課用圖_bad.csv", "sha256": "0" * 64, "source_decision_id": "D"}]})
    write_json(bundle / "approval_manifest.json", {"status": "PASS", "decisions": [{"decision_id": "D", "human_decision": "HUMAN_APPROVED"}]})
    write_json(bundle / "provenance_map.json", {"rows": []})
    write_json(bundle / "visible_file_manifest.json", {"files": []})
    write_json(bundle / "verification_results.json", {"status": "PENDING_VERIFICATION", "errors": ["BUNDLE_NOT_VERIFIED"], "warnings": []})
    write_json(bundle / "bundle_hashes.json", {"files": []})
    leak = bundle / "sanitized_inputs" / "01_生管課用圖_bad.csv"
    leak.parent.mkdir(parents=True, exist_ok=True)
    leak.write_text("C:\\Users\\alex1\\OneDrive\\Desktop\\All Projects\\secret", encoding="utf-8")
    result = subprocess.run([PY, "scripts/verify_generator_bundle.py", "--bundle-dir", str(bundle), "--output", str(bundle / "verification.json")], cwd=ROOT)
    assert_true(result.returncode != 0, "bundle verifier must reject source path and target-output leaks")


def create_reference_v3_pdf(path: Path, page_texts: list[str]) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    for index, text in enumerate(page_texts, start=1):
        c.setFont("Helvetica", 14)
        c.drawString(72, height - 72, text)
        c.setFont("Helvetica", 8)
        c.drawString(width - 240, 48, f"TITLE BLOCK {index}")
        c.showPage()
    c.save()


def create_reference_v4_image_only_pdf(path: Path, page_texts: list[str]) -> None:
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    image_dir = path.parent / f"{path.stem}_images"
    image_dir.mkdir(parents=True, exist_ok=True)
    try:
        font = ImageFont.truetype("arial.ttf", 64)
    except Exception:
        font = ImageFont.load_default()
    for index, text in enumerate(page_texts, start=1):
        image = Image.new("RGB", (1100, 420), "white")
        draw = ImageDraw.Draw(image)
        draw.text((60, 70), "PROJECT 1999001", fill="black", font=font)
        draw.text((60, 210), text, fill="black", font=font)
        png = image_dir / f"page_{index}.png"
        image.save(png)
        c.drawImage(str(png), 60, height / 2, width=500, height=190)
        c.showPage()
    c.save()
    shutil.rmtree(image_dir)


def create_reference_v4_layout_prior_pdf(path: Path, page_texts: list[str]) -> None:
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    image_dir = path.parent / f"{path.stem}_images"
    image_dir.mkdir(parents=True, exist_ok=True)
    try:
        font = ImageFont.truetype("arial.ttf", 42)
        small = ImageFont.truetype("arial.ttf", 26)
    except Exception:
        font = ImageFont.load_default()
        small = ImageFont.load_default()
    for index, text in enumerate(page_texts, start=1):
        image = Image.new("RGB", (850, 1100), "white")
        draw = ImageDraw.Draw(image)
        for x in range(80, 760, 120):
            draw.line((x, 120, x, 760), fill="black", width=3)
        for y in range(120, 760, 90):
            draw.line((80, y, 760, y), fill="black", width=3)
        draw.rectangle((450, 790, 820, 1060), outline="black", width=4)
        for y in range(840, 1040, 50):
            draw.line((450, y, 820, y), fill="black", width=2)
        draw.text((110, 60), "PROJECT 1999001", fill="black", font=font)
        draw.text((480, 805), text, fill="black", font=small)
        png = image_dir / f"page_{index}.png"
        image.save(png)
        c.drawImage(str(png), 0, 0, width=width, height=height)
        c.showPage()
    c.save()
    shutil.rmtree(image_dir)


def create_blank_reference_pdf(path: Path) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    c.showPage()
    c.save()


def reference_v3_manifest(path: Path, project_id: str, files: list[dict[str, str]]) -> None:
    rows = []
    for item in files:
        pdf = Path(item["path"])
        rows.append({
            "project_id": project_id,
            "neutral_reference_file_id": item["file_id"],
            "path": str(pdf),
            "sha256": __import__("hashlib").sha256(pdf.read_bytes()).hexdigest().upper(),
            "role_hint": item.get("role_hint", "supporting_document_non_generator"),
            "metadata_text": item.get("metadata_text", pdf.name),
        })
    write_json(path, {"project_id": project_id, "files": rows})


def test_reference_detection_v3_boundary_and_page_level_sets() -> None:
    project_id = "1999001"
    work = ROOT / "tmp" / "reference_detection_v3_tests" / "boundary"
    if work.exists():
        shutil.rmtree(work)
    fixture = read_json(ROOT / "evals" / "fixtures" / "reference_detection_v3" / "fixture_manifest.json")
    cases = {case["case_id"]: case for case in fixture["cases"]}

    production = work / "forbidden_production_control" / "synthetic_production.pdf"
    electrical = work / "electrical" / "synthetic_electrical.pdf"
    combined = work / "combined" / "synthetic_combined.pdf"
    create_reference_v3_pdf(production, cases["target_under_production_control"]["pages"])
    create_reference_v3_pdf(electrical, cases["ordinary_electrical_under_electrical"]["pages"])
    create_reference_v3_pdf(combined, cases["combined_all_three"]["pages"])

    candidate_manifest = work / "candidate_manifest.json"
    reference_v3_manifest(candidate_manifest, project_id, [
        {"file_id": "REF-PROD-FORBIDDEN-FOLDER", "path": str(production), "role_hint": "forbidden_production_control_file", "metadata_text": "production-control-folder"},
        {"file_id": "REF-ELECTRICAL", "path": str(electrical), "role_hint": "forbidden_electrical_drawing", "metadata_text": "electrical-folder"},
        {"file_id": "REF-COMBINED-CJK-合併", "path": str(combined), "role_hint": "supporting_document_non_generator", "metadata_text": "合併 package"},
    ])
    output_dir = work / "out"
    run([PY, "scripts/detect_reference_presence_v3.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(output_dir), "--task-id", "TEST-REFDET-V3-BOUNDARY"])
    run([PY, "scripts/verify_reference_detection_output.py", "--output-dir", str(output_dir)])

    effective = read_json(output_dir / "effective_reference_set.json")
    pages = read_json(output_dir / "reference_page_classifications.json")["page_classifications"]
    audit = read_json(output_dir / "reference_detection_audit.json")
    assert_true(effective["status"] == "VERIFIED_ALL_THREE_COMBINED_PACKAGE", "combined package with all three page types must verify")
    assert_true(audit["temporary_workspace_removed"], "temporary reference renders must be deleted")
    assert_true(any(row["neutral_reference_file_id"] == "REF-PROD-FORBIDDEN-FOLDER" and row["page_classification"] == "PRODUCTION_DRAWING" for row in pages), "target under production-control folder must qualify as reviewer evidence by content")
    assert_true(any(row["neutral_reference_file_id"] == "REF-ELECTRICAL" and row["page_classification"] == "ELECTRICAL_DRAWING" for row in pages), "ordinary electrical drawing must stay non-target")

    source_manifest = work / "source_manifest.json"
    sheet_adjudication = work / "sheet_adjudication.json"
    generator_manifest = work / "generator_input_manifest.json"
    write_json(source_manifest, {
        "manifest_id": "ref-v3-boundary-source",
        "files": [
            {
                "file_id": "REF-PROD-FORBIDDEN-FOLDER",
                "relative_path": "forbidden_production_control/synthetic_production.pdf",
                "file_name": "synthetic_production.pdf",
                "primary_role": "forbidden_production_control_file",
                "generator_eligibility": "FORBIDDEN",
                "absolute_path": str(production),
            }
        ],
        "worksheets": [],
    })
    write_json(sheet_adjudication, {"adjudication_id": "ref-v3-boundary-sheets", "worksheets": []})
    run([PY, "scripts/build_generator_bundle.py", "--source-manifest", str(source_manifest), "--sheet-adjudication", str(sheet_adjudication), "--bundle-dir", str(work / "generator_bundle"), "--output-manifest", str(generator_manifest), "--run-id", "REF-V3-BOUNDARY", "--project-id", project_id])
    generator = read_json(generator_manifest)
    assert_true(generator["allowed_files"] == [], "reviewer-reference PDF must not enter generator bundle")
    assert_true("REF-PROD-FORBIDDEN-FOLDER" not in json.dumps(generator, ensure_ascii=False), "generator manifest must not expose reference file IDs")


def test_reference_detection_v3_duplicates_identity_and_missing_types() -> None:
    project_id = "1999001"
    work = ROOT / "tmp" / "reference_detection_v3_tests" / "duplicates"
    if work.exists():
        shutil.rmtree(work)
    fixture = read_json(ROOT / "evals" / "fixtures" / "reference_detection_v3" / "fixture_manifest.json")
    cases = {case["case_id"]: case for case in fixture["cases"]}

    production = work / "production.pdf"
    duplicate = work / "duplicate_named_punch.pdf"
    sheetmetal = work / "sheetmetal.pdf"
    wrong_project = work / "wrong_project_punch.pdf"
    near_duplicate = work / "production_near_duplicate.pdf"
    create_reference_v3_pdf(production, cases["missing_third_with_duplicate"]["pages"])
    shutil.copy2(production, duplicate)
    create_reference_v3_pdf(sheetmetal, cases["sheetmetal_only"]["pages"])
    create_reference_v3_pdf(wrong_project, cases["mismatched_project_identity"]["pages"])
    create_reference_v3_pdf(near_duplicate, ["PROJECT 1999001 PRODUCTION_DRAWING NEAR DUPLICATE"])

    candidate_manifest = work / "candidate_manifest.json"
    reference_v3_manifest(candidate_manifest, project_id, [
        {"file_id": "REF-PROD", "path": str(production), "metadata_text": "production"},
        {"file_id": "REF-DUP-PUNCH-NAME", "path": str(duplicate), "metadata_text": "punch filename but exact duplicate"},
        {"file_id": "REF-SHEETMETAL", "path": str(sheetmetal), "metadata_text": "sheetmetal"},
        {"file_id": "REF-WRONG-PROJECT", "path": str(wrong_project), "metadata_text": "punch wrong project"},
        {"file_id": "REF-NEAR-DUP-PROD", "path": str(near_duplicate), "metadata_text": "production near duplicate"},
    ])
    output_dir = work / "out"
    run([PY, "scripts/detect_reference_presence_v3.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(output_dir), "--task-id", "TEST-REFDET-V3-DUP"])
    run([PY, "scripts/verify_reference_detection_output.py", "--output-dir", str(output_dir)])
    effective = read_json(output_dir / "effective_reference_set.json")
    docs = read_json(output_dir / "reference_document_classifications.json")["document_classifications"]
    pages = read_json(output_dir / "reference_page_classifications.json")["page_classifications"]
    assert_true(effective["status"] == "PARTIAL_REFERENCE_SET", "duplicate or wrong-project punch must not create false all-three set")
    assert_true(set(effective["target_types_present"]) == {"PRODUCTION_DRAWING", "SHEETMETAL_DRAWING"}, "only confirmed production and sheetmetal should count")
    assert_true(any(doc["neutral_reference_file_id"] == "REF-DUP-PUNCH-NAME" and doc["duplicate_of"] == "REF-PROD" for doc in docs), "exact duplicate must be recorded")
    assert_true(any(row["neutral_reference_file_id"] == "REF-WRONG-PROJECT" and row["project_identity_status"] == "CONFLICT" for row in pages), "mismatched project identity must be marked conflict")
    assert_true(any(row["neutral_reference_file_id"] == "REF-NEAR-DUP-PROD" and row["page_classification"] == "PRODUCTION_DRAWING" for row in pages), "near duplicate should remain same target type, not a missing third output")


def test_reference_detector_v3_known_positive_recall_gate() -> None:
    summary_path = ROOT / "reports" / "baseline-024" / "reference-detector-calibration" / "known_positive_replay_summary.json"
    summary = read_json(summary_path)
    expected_projects = {
        "1110101",
        "1110103",
        "1110104",
        "1110203",
        "1110204",
        "1110205",
        "1110405",
        "1110410",
        "1110704",
        "1110801",
        "1120207",
        "1120305",
        "1120308",
    }
    project_results = {row["project_id"]: row for row in summary["project_results"]}
    assert_true(summary["detector_version"] == "target_output_detection_v3_page_content_isolated", "unexpected detector under v3 replay gate")
    assert_true(summary["status"] == "DETECTOR_V3_RECALL_FAIL", "v3 known-positive replay must remain a failed gate until superseded")
    assert_true(set(summary["known_positive_projects"]) == expected_projects, "known-positive control set changed")
    assert_true(set(project_results) == expected_projects, "known-positive replay result set changed")
    assert_true(summary["positive_projects_detected_all_three"] == 0, "v3 must not be recorded as detecting any all-three known positive")
    assert_true(summary["false_negative_output_type_count"] == 31, "v3 false-negative count changed without a new detector version")
    assert_true(summary["project_identity_mismatch_count"] == 0, "project identity mismatch count changed")
    assert_true(summary["per_type_recall"]["PRODUCTION_DRAWING"]["detected"] == 0, "v3 production recall unexpectedly changed")
    assert_true(summary["per_type_recall"]["SHEETMETAL_DRAWING"]["detected"] == 8, "v3 sheetmetal recall unexpectedly changed")
    assert_true(summary["per_type_recall"]["PUNCH_DRAWING"]["detected"] == 0, "v3 punch recall unexpectedly changed")
    assert_true(summary["actual_vision_agents"]["image_only_pages_inspected_by_actual_vision_agents"] == 0, "local v3 replay must not masquerade as actual vision classification")
    assert_true(summary["classifier_models"][0]["actual_model"] == "local_poppler_pypdf_deterministic_reference_detector_v3", "v3 replay model provenance changed")
    for project_id in expected_projects:
        result = project_results[project_id]
        assert_true(result["missed_output_types"], f"{project_id} must remain a recorded v3 missed-positive regression case")
        assert_true("IMAGE_OR_NO_TARGET_TEXT_WITHOUT_REAL_VISION_CLASSIFICATION" in result["failure_reason"], f"{project_id} missing real-vision failure reason")


def test_windows_media_ocr_probe_minimized() -> None:
    probe = read_json(ROOT / "reports" / "baseline-024" / "reference-detector-calibration" / "windows_media_ocr_local_probe.json")
    assert_true(probe["synthetic_execution"]["status"] == "PASS", "Windows.Media.Ocr synthetic execution must pass")
    assert_true("PUNCH_DRAWING" in probe["synthetic_execution"]["role_hits"], "synthetic OCR must return a minimized role code")
    assert_true(probe["language_availability"]["english_available"] is True, "English OCR language should be available")
    assert_true(probe["language_availability"]["traditional_chinese_available"] is True, "Traditional Chinese OCR language should be available")
    assert_true(probe["language_availability"]["simplified_chinese_available"] is False, "Simplified Chinese availability should be recorded as false in this local probe")
    assert_true(probe["privacy"]["private_content_transmitted_outside_machine"] == 0, "OCR probe must not transmit private content")
    assert_true(probe["privacy"]["ocr_output_appeared_in_stdout"] is False, "OCR output must not appear in stdout")
    assert_true(probe["private_page_execution"]["status"] == "SKIPPED_NETWORK_DISABLE_BOUNDARY_NOT_ENFORCED", "private-page probe must stay skipped without enforceable network disable")


def test_reference_detection_v4_multisignal_and_conflicts() -> None:
    project_id = "1999001"
    work = ROOT / "tmp" / "reference_detection_v4_tests" / "multisignal"
    if work.exists():
        shutil.rmtree(work)
    combined = work / "combined.pdf"
    electrical = work / "misleading_production_name.pdf"
    source_confuser = work / "source_confuser.pdf"
    create_reference_v3_pdf(combined, [
        "PROJECT 1999001 PRODUCTION DRAWING",
        "PROJECT 1999001 SHEET METAL DRAWING",
        "PROJECT 1999001 PUNCH DRAWING",
    ])
    create_reference_v3_pdf(electrical, ["PROJECT 1999001 ELECTRICAL DRAWING"])
    create_reference_v3_pdf(source_confuser, ["PROJECT 1999001 SOURCE DOCUMENT PUNCH DRAWING"])
    candidate_manifest = work / "candidate_manifest.json"
    reference_v3_manifest(candidate_manifest, project_id, [
        {"file_id": "REF-V4-COMBINED", "path": str(combined), "metadata_text": "combined target package"},
        {"file_id": "REF-V4-ELECTRICAL-MISLEADING", "path": str(electrical), "role_hint": "completed_production_drawing_reference", "metadata_text": "PRODUCTION_DRAWING misleading filename"},
        {"file_id": "REF-V4-SOURCE-CONFUSER", "path": str(source_confuser), "metadata_text": "source document mentions punch"},
    ])
    output_dir = work / "out"
    run([PY, "scripts/detect_reference_presence_v4.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(output_dir), "--task-id", "TEST-REFDET-V4-MULTI", "--disable-ocr"])
    run([PY, "scripts/verify_reference_detection_v4_output.py", "--output-dir", str(output_dir)])
    effective = read_json(output_dir / "effective_reference_set.json")
    pages = read_json(output_dir / "reference_page_classifications.json")["page_classifications"]
    audit = read_json(output_dir / "reference_detection_audit.json")
    assert_true(effective["status"] == "VERIFIED_ALL_THREE_COMBINED_PACKAGE", "v4 must segment combined target packages")
    assert_true(set(effective["target_types_present"]) == {"PRODUCTION_DRAWING", "SHEETMETAL_DRAWING", "PUNCH_DRAWING"}, "v4 must detect all target roles")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-ELECTRICAL-MISLEADING" and row["page_classification"] == "ELECTRICAL_DRAWING" for row in pages), "explicit electrical page content must override misleading target hints")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-SOURCE-CONFUSER" and row["page_classification"] == "AMBIGUOUS" for row in pages), "source document with target words must fail closed")
    assert_true(audit["temporary_workspace_removed"], "v4 temporary renders must be deleted")
    assert_true(audit["generator_isolation_pass"] and audit["source_review_blindness_pass"], "v4 must preserve downstream isolation")


def test_reference_detection_v4_image_only_ocr_and_failure_modes() -> None:
    project_id = "1999001"
    work = ROOT / "tmp" / "reference_detection_v4_tests" / "image_ocr"
    if work.exists():
        shutil.rmtree(work)
    production = work / "production.pdf"
    sheetmetal = work / "sheetmetal.pdf"
    image_punch = work / "image_punch.pdf"
    create_reference_v3_pdf(production, ["PROJECT 1999001 PRODUCTION DRAWING"])
    create_reference_v3_pdf(sheetmetal, ["PROJECT 1999001 SHEET METAL DRAWING"])
    create_reference_v4_image_only_pdf(image_punch, ["PUNCH DRAWING"])
    candidate_manifest = work / "candidate_manifest.json"
    reference_v3_manifest(candidate_manifest, project_id, [
        {"file_id": "REF-V4-PRODUCTION", "path": str(production), "metadata_text": "production"},
        {"file_id": "REF-V4-SHEETMETAL", "path": str(sheetmetal), "metadata_text": "sheetmetal"},
        {"file_id": "REF-V4-IMAGE-PUNCH", "path": str(image_punch), "metadata_text": "image only punch"},
    ])
    output_dir = work / "out"
    run([PY, "scripts/detect_reference_presence_v4.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(output_dir), "--task-id", "TEST-REFDET-V4-OCR"])
    run([PY, "scripts/verify_reference_detection_v4_output.py", "--output-dir", str(output_dir)])
    effective = read_json(output_dir / "effective_reference_set.json")
    pages = read_json(output_dir / "reference_page_classifications.json")["page_classifications"]
    assert_true(effective["status"] == "VERIFIED_ALL_THREE_BY_CONTENT", "OCR-backed image-only punch page should complete all-three set")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-IMAGE-PUNCH" and row["page_classification"] == "PUNCH_DRAWING" and "WINDOWS_MEDIA_OCR" in row["evidence_channel_codes"] for row in pages), "image-only target page must use Windows OCR signal")

    fail_dir = work / "out_ocr_fail"
    run([PY, "scripts/detect_reference_presence_v4.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(fail_dir), "--task-id", "TEST-REFDET-V4-OCR-FAIL", "--simulate-ocr-failure"])
    run([PY, "scripts/verify_reference_detection_v4_output.py", "--output-dir", str(fail_dir)])
    fail_pages = read_json(fail_dir / "reference_page_classifications.json")["page_classifications"]
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-IMAGE-PUNCH" and row["page_classification"] == "UNCLASSIFIED" and "OCR_UNAVAILABLE" in row["evidence_channel_codes"] for row in fail_pages), "OCR failure must fail closed for image-only target pages")

    missing_lang_dir = work / "out_missing_lang"
    run([PY, "scripts/detect_reference_presence_v4.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(missing_lang_dir), "--task-id", "TEST-REFDET-V4-MISSING-LANG", "--ocr-language", "zh-Hans-CN"])
    run([PY, "scripts/verify_reference_detection_v4_output.py", "--output-dir", str(missing_lang_dir)])
    missing_lang_pages = read_json(missing_lang_dir / "reference_page_classifications.json")["page_classifications"]
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-IMAGE-PUNCH" and row["page_classification"] == "UNCLASSIFIED" for row in missing_lang_pages), "missing OCR language support must fail closed")


def test_reference_detection_v4_private_ocr_disabled_layout_prior() -> None:
    project_id = "1999001"
    work = ROOT / "tmp" / "reference_detection_v4_tests" / "layout_prior"
    if work.exists():
        shutil.rmtree(work)
    production = work / "image_production.pdf"
    sheetmetal = work / "image_sheetmetal.pdf"
    punch = work / "image_punch.pdf"
    blank_prior = work / "blank_role_prior.pdf"
    create_reference_v4_layout_prior_pdf(production, ["CONTROL PANEL"])
    create_reference_v4_layout_prior_pdf(sheetmetal, ["SHEET LAYOUT"])
    create_reference_v4_layout_prior_pdf(punch, ["HOLE LAYOUT"])
    create_blank_reference_pdf(blank_prior)
    candidate_manifest = work / "candidate_manifest.json"
    reference_v3_manifest(candidate_manifest, project_id, [
        {"file_id": "REF-V4-LAYOUT-PROD", "path": str(production), "role_hint": "completed_production_drawing_reference", "metadata_text": "1999001 production role prior"},
        {"file_id": "REF-V4-LAYOUT-SHEET", "path": str(sheetmetal), "role_hint": "completed_sheetmetal_drawing_reference", "metadata_text": "1999001 sheetmetal role prior"},
        {"file_id": "REF-V4-LAYOUT-PUNCH", "path": str(punch), "role_hint": "completed_punch_drawing_reference", "metadata_text": "1999001 punch role prior"},
        {"file_id": "REF-V4-BLANK-PRIOR", "path": str(blank_prior), "role_hint": "completed_production_drawing_reference", "metadata_text": "1999001 blank role prior"},
    ])
    output_dir = work / "out"
    run([PY, "scripts/detect_reference_presence_v4.py", "--project-id", project_id, "--candidate-manifest", str(candidate_manifest), "--output-dir", str(output_dir), "--task-id", "TEST-REFDET-V4-LAYOUT-PRIOR", "--disable-ocr"])
    run([PY, "scripts/verify_reference_detection_v4_output.py", "--output-dir", str(output_dir)])
    effective = read_json(output_dir / "effective_reference_set.json")
    pages = read_json(output_dir / "reference_page_classifications.json")["page_classifications"]
    assert_true(effective["status"] == "VERIFIED_ALL_THREE_BY_CONTENT", "layout-confirmed weak priors should recover OCR-disabled image-only target pages")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-LAYOUT-PROD" and row["page_classification"] == "PRODUCTION_DRAWING" and row["classification_method"] == "LAYOUT_CONFIRMED_WEAK_ROLE_PRIOR_NO_TEXT" for row in pages), "production role prior must require layout confirmation")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-LAYOUT-SHEET" and row["page_classification"] == "SHEETMETAL_DRAWING" for row in pages), "sheetmetal role prior must recover only with layout confirmation")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-LAYOUT-PUNCH" and row["page_classification"] == "PUNCH_DRAWING" for row in pages), "punch role prior must recover only with layout confirmation")
    assert_true(any(row["neutral_reference_file_id"] == "REF-V4-BLANK-PRIOR" and row["page_classification"] == "UNCLASSIFIED" for row in pages), "weak role prior without layout evidence must stay unclassified")
    assert_true(not any("WINDOWS_MEDIA_OCR" in row.get("evidence_channel_codes", []) for row in pages), "private OCR disabled fixture must not use OCR")


def test_qualification_recovery_controller_state() -> None:
    work = ROOT / "tmp" / "qualification_recovery_controller_test"
    if work.exists():
        shutil.rmtree(work)
    report_dir = work / "reports"
    queue_path = work / "queue.json"
    run([
        PY,
        "scripts/run_qualification_recovery.py",
        "--report-dir",
        str(report_dir),
        "--queue-path",
        str(queue_path),
        "--skip-capability-probe",
    ])
    state_path = report_dir / "recovery_state.json"
    state = read_json(state_path)
    errors = validate(state, read_json(ROOT / "schemas/qualification_recovery_state.schema.json"))
    assert_true(not errors, f"qualification recovery state schema errors: {errors}")
    assert_true(state["decision_id"] == "D-0021", "recovery state must bind to accepted decision D-0021")
    assert_true(state["current_allowed_eval"]["current_count"] == 13, "current ALLOWED_EVAL count must remain 13")
    assert_true(state["current_allowed_eval"]["deficit"] == 11, "deficit must remain 11")
    assert_true(state["privacy"]["approval_status"] == "NOT_APPROVED", "privacy approval must remain NOT_APPROVED")
    assert_true(state["privacy"]["private_reference_pages_inspected_by_actual_vision_agents"] == 0, "private vision inspection count must stay zero")
    assert_true(state["detector_calibration"]["regression_gate_behavior_status"] == "PASS_BLOCKS_KNOWN_FAILING_DETECTOR", "gate behavior must be reported separately")
    assert_true(state["detector_calibration"]["detector_performance_status"] == "FAIL", "v3 detector performance must remain failed")
    assert_true("RECOVERABLE_DETECTOR_LIMITATION" in state["blocker_taxonomy"], "recoverable detector blocker class missing")
    assert_true(state["next_selected_action"]["action_id"] == "RUN_LOCAL_CAPABILITY_DISCOVERY", "test-mode next action should be capability discovery")
    assert_true(queue_path.exists(), "qualification recovery queue must be written")
    queue = read_json(queue_path)
    assert_true(queue["privacy_minimization"]["stores_private_paths"] is False, "queue must not store private paths")
    assert_true(queue["privacy_minimization"]["stores_reference_content"] is False, "queue must not store reference content")


def test_sheetmetal_v1_modular_foundation() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_pipeline_test"
    if work.exists():
        shutil.rmtree(work)
    fixture = ROOT / "evals" / "fixtures" / "sheetmetal-v1" / "complete_pipeline_fixture.json"
    run([PY, "scripts/sheetmetal_v1.py", "--fixture", str(fixture), "--output-dir", str(work)])
    fixture_data = read_json(fixture)
    for fact in fixture_data["source_facts"]:
        errors = validate(fact, read_json(ROOT / "schemas/source_fact.schema.json"))
        assert_true(not errors, f"source fact schema errors: {errors}")
    for panel_definition in fixture_data["panel_definitions"]:
        errors = validate(panel_definition, read_json(ROOT / "schemas/panel_definition.schema.json"))
        assert_true(not errors, f"panel definition schema errors: {errors}")

    schema_pairs = {
        "source_evidence.json": "schemas/source_evidence.schema.json",
        "component_register.json": "schemas/component_register.schema.json",
        "panel_assignment.json": "schemas/panel_assignment.schema.json",
        "panel_graph.json": "schemas/panel_graph.schema.json",
        "panel_constraint_model.json": "schemas/panel_constraint_model.schema.json",
        "sheetmetal_drawing_model.json": "schemas/sheetmetal_drawing_model.schema.json",
        "drawing_provenance_map.json": "schemas/drawing_provenance_map.schema.json",
    }
    for instance, schema in schema_pairs.items():
        errors = validate_file(work / instance, ROOT / schema)
        assert_true(not errors, f"{instance} failed {schema}: {errors}")

    register = read_json(work / "component_register.json")
    for component_type in register["component_types"]:
        errors = validate(component_type, read_json(ROOT / "schemas/component_type.schema.json"))
        assert_true(not errors, f"component type schema errors: {errors}")
        if component_type.get("geometry"):
            errors = validate(component_type["geometry"], read_json(ROOT / "schemas/component_geometry.schema.json"))
            assert_true(not errors, f"component geometry schema errors: {errors}")
    for component_instance in register["component_instances"]:
        errors = validate(component_instance, read_json(ROOT / "schemas/component_instance.schema.json"))
        assert_true(not errors, f"component instance schema errors: {errors}")
    rule_pack = read_json(ROOT / "rules" / "sheetmetal-v1" / "accessory_rules.synthetic.json")
    for rule in rule_pack["rules"]:
        errors = validate(rule, read_json(ROOT / "schemas/accessory_rule.schema.json"))
        assert_true(not errors, f"accessory rule schema errors: {errors}")
    instances = {row["component_key"]: row for row in register["component_instances"]}
    fan_qty = instances["fan_main"]["quantity"]
    assert_true(fan_qty["required_qty"] == 1, "required quantity must come from requirement evidence")
    assert_true(fan_qty["ordered_qty"] == 2, "procurement quantity must remain separate")

    breaker = instances["breaker_conflict"]
    assert_true(breaker["status"] == "CONFLICT", "conflicting component model must not be resolved")
    assert_true(any(row["field"] == "model" for row in breaker["conflict_records"]), "model conflict record missing")

    accessories = read_json(work / "accessory_requirements.json")
    assert_true(accessories["duplicate_accessory_count"] == 0, "explicit accessory must prevent duplicate inferred accessory")
    assert_true(accessories["requirements"][0]["status"] == "SATISFIED_BY_EXPLICIT_SOURCE", "explicit accessory should satisfy rule output")
    assert_true(not accessories["generated_component_instances"], "rule must not create duplicate explicit accessory instance")

    assignment = read_json(work / "panel_assignment.json")
    assert_true(any(row["reason"] == "POST_DESIGN_OR_REFERENCE_ASSIGNMENT_REJECTED" for row in assignment["rejected_assignments"]), "post-design allocation label must be rejected")
    assert_true(all(row["placement_resolved"] is False for row in assignment["assignments"]), "panel assignment must stay separate from placement")

    graph = read_json(work / "panel_graph.json")
    functional_edges = [row for row in graph["edges"] if row["edge_type"] in {"CONNECTS_TO", "SUPPLIES", "PROTECTS", "CONTROLS", "MEASURES", "REPORTS_TO", "INTERLOCKS_WITH"}]
    assert_true(functional_edges, "inventory-only mode should record missing functional relationships as unverified")
    assert_true(all(row["status"] == "UNVERIFIED" for row in functional_edges), "inventory-only mode must not invent supported functional edges")
    fan_id = instances["fan_main"]["component_instance_id"]
    fan_edge_types = {row["edge_type"] for row in graph["edges"] if row["from_node_id"] == f"CINST:{fan_id}"}
    assert_true({"REQUIRED_BY", "ASSIGNED_TO_PANEL", "REQUIRES_ACCESSORY", "REQUIRES_CUTOUT", "MOUNTED_ON"}.issubset(fan_edge_types), "component should support several typed relationships")

    drawing = read_json(work / "sheetmetal_drawing_model.json")
    dims = drawing["dimensions"][0]
    assert_true(dims["width_mm"] == 800 and dims["height_mm"] == 2300 and dims["depth_mm"] == 800, "width, height, and depth must remain individually named and not reordered")
    assert_true(drawing["panel_topology"]["rejected_raw_dimension_strings"], "ambiguous raw dimension string must not be silently parsed")

    constraints = read_json(work / "panel_constraint_model.json")
    assert_true(constraints["placements"], "valid placement should be recorded separately")
    assert_true(any("CONTAINMENT" in row["unsatisfied_constraints"] and row["soft_objective_override_rejected"] for row in constraints["failed_placements"]), "hard placement constraints cannot be overridden by soft objectives")

    provenance = read_json(work / "drawing_provenance_map.json")
    assert_true(provenance["coverage_status"] == "PASS", "every critical drawing-model fact must have support or safe unresolved status")

    validation = read_json(work / "validation_report.json")
    assert_true(validation["status"] == "PASS", f"sheetmetal v1 validation failed: {validation}")
    assert_true(validation["leakage_scan"]["status"] == "PASS", "completed-reference IDs and content must not enter generator artifacts")
    serialized_generator_artifacts = "\n".join((work / name).read_text(encoding="utf-8") for name in [
        "component_register.json",
        "panel_assignment.json",
        "panel_graph.json",
        "accessory_requirements.json",
        "panel_topology.json",
        "panel_constraint_model.json",
        "sheetmetal_drawing_model.json",
        "drawing_provenance_map.json",
    ])
    for forbidden in ["REF-COMPLETE-001", "completed-reference-secret", "post-design-secret"]:
        assert_true(forbidden not in serialized_generator_artifacts, f"forbidden token leaked to generator artifact: {forbidden}")


def test_sheetmetal_v1_private_workspace_boundary() -> None:
    import verify_frozen_workflow as verifier

    private_probe = ".private/sheetmetal-v1/1110101/private-output-probe.json"
    ignored = subprocess.run(
        [verifier.git_executable(), "check-ignore", private_probe],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )
    assert_true(ignored.returncode == 0, f"private workspace path must be ignored: {ignored.stderr or ignored.stdout}")

    tracked = subprocess.run(
        [verifier.git_executable(), "ls-files", "--", ".private"],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    assert_true(not tracked.stdout.strip(), "no private workspace artifact may be tracked")


def test_sheetmetal_v1_source_fact_extractor() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_source_fact_extractor_test"
    if work.exists():
        shutil.rmtree(work)
    bundle = work / "bundle"
    inputs = bundle / "sanitized_inputs"
    out = work / "out"
    inputs.mkdir(parents=True)
    secret_model = "MODEL-SECRET-123"
    secret_customer = "CUSTOMER-SECRET"
    import csv

    material_csv = inputs / "material.csv"
    with material_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["item", "model", "required_qty", "customer_note"])
        writer.writeheader()
        writer.writerow({"item": "M1", "model": secret_model, "required_qty": "2", "customer_note": secret_customer})
    procurement_csv = inputs / "procurement.csv"
    with procurement_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["item", "model", "qty"])
        writer.writeheader()
        writer.writerow({"item": "M1", "model": secret_model, "qty": "5"})
    generic_csv = inputs / "generic_current_input.csv"
    with generic_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["item", "model", "qty"])
        writer.writeheader()
        writer.writerow({"item": "M2", "model": "GENERIC-CURRENT-INPUT", "qty": "1"})

    material_sha = __import__("hashlib").sha256(material_csv.read_bytes()).hexdigest().upper()
    procurement_sha = __import__("hashlib").sha256(procurement_csv.read_bytes()).hexdigest().upper()
    generic_sha = __import__("hashlib").sha256(generic_csv.read_bytes()).hexdigest().upper()
    write_json(bundle / "bundle_manifest.json", {
        "bundle_id": "SYNTH-SOURCE-FACT-BUNDLE",
        "project_id": "SYNTH-SMV1-SOURCE",
        "selection_id": "SYNTH-SELECTION",
        "status": "BUILT",
        "artifacts": [
            {"artifact_id": "ART-MAT", "path": "sanitized_inputs/material.csv", "sha256": material_sha, "source_decision_id": "DEC-MAT"},
            {"artifact_id": "ART-PO", "path": "sanitized_inputs/procurement.csv", "sha256": procurement_sha, "source_decision_id": "DEC-PO"},
            {"artifact_id": "ART-GEN", "path": "sanitized_inputs/generic_current_input.csv", "sha256": generic_sha, "source_decision_id": "DEC-GEN"},
        ],
    })
    write_json(bundle / "provenance_map.json", {
        "bundle_id": "SYNTH-SOURCE-FACT-BUNDLE",
        "selection_id": "SYNTH-SELECTION",
        "rows": [
            {"neutral_source_id": "SRC-MAT", "sanitized_artifact": "sanitized_inputs/material.csv", "source_decision_id": "DEC-MAT", "source_file_id": "FILE-MAT", "source_sheet_id": "SHEET-MAT"},
            {"neutral_source_id": "SRC-PO", "sanitized_artifact": "sanitized_inputs/procurement.csv", "source_decision_id": "DEC-PO", "source_file_id": "FILE-PO", "source_sheet_id": "SHEET-PO"},
            {"neutral_source_id": "SRC-GEN", "sanitized_artifact": "sanitized_inputs/generic_current_input.csv", "source_decision_id": "DEC-GEN", "source_file_id": "FILE-GEN", "source_sheet_id": "SHEET-GEN"},
        ],
    })
    classification = work / "classification.json"
    write_json(classification, {
        "status": "SOURCE_ROLE_CHRONOLOGY_CLASSIFICATION_PASS",
        "approved_eval_items": [
            {
                "decision_id": "DEC-MAT",
                "source_role_classification": "MATERIAL_REQUIREMENT",
                "chronology_classification": "PRE_DESIGN",
                "completed_reference_or_derivative": False,
            },
            {
                "decision_id": "DEC-PO",
                "source_role_classification": "PROCUREMENT_EVIDENCE",
                "chronology_classification": "PRE_DESIGN",
                "completed_reference_or_derivative": False,
            },
            {
                "decision_id": "DEC-GEN",
                "source_role_classification": "PERMITTED_CURRENT_PROJECT_INPUT_ROLE",
                "chronology_classification": "CURRENT_PROJECT_PRE_OR_DURING_DESIGN_INPUT_NOT_PRODUCTION_APPROVED",
                "completed_reference_or_derivative": "NO_SIGNAL_IN_APPROVED_METADATA",
            },
        ],
    })
    result = subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--bundle-dir",
            str(bundle),
            "--source-classification",
            str(classification),
            "--output-dir",
            str(out),
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    assert_true(secret_model not in result.stdout and secret_customer not in result.stdout, "source values must not be printed to stdout")
    assert_true(secret_model not in result.stderr and secret_customer not in result.stderr, "source values must not be printed to stderr")
    errors = validate_file(out / "source_fact_model.json", ROOT / "schemas/source_fact_model.schema.json")
    assert_true(not errors, f"source fact model schema errors: {errors}")
    model = read_json(out / "source_fact_model.json")
    assert_true(model["validation"]["status"] == "PASS", "source fact extraction must pass")
    assert_true(model["validation"]["source_line_count"] == 3, "every source row must be accounted for")
    assert_true(model["validation"]["silently_discarded_authorized_source_lines"] == 0, "authorized source lines must not disappear")
    assert_true(model["quantity_stage_counts"]["required_qty"] == 2, "material and generic current quantities should be required quantities")
    assert_true(model["quantity_stage_counts"]["ordered_qty"] == 1, "procurement quantity should be ordered quantity")
    assert_true(all(row["status"] == "REPRESENTED" for row in model["source_line_accounting"]), "all synthetic source rows should be represented")
    assert_true(model["validation"]["private_content_transmission_count"] == 0, "private transmission count must stay zero")


def test_sheetmetal_v1_component_register_from_source_facts() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_component_register_from_source_facts_test"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)
    out = work / "out"
    secret_model = "REGISTER-SECRET-MODEL"
    secret_maker = "REGISTER-SECRET-MAKER"
    source_model = {
        "schema_version": "sheetmetal-v1.source_fact_model.v1",
        "project_id": "SYNTH-SMV1-REGISTER",
        "source_mode": "SOURCE_MODE_A_INVENTORY_ONLY",
        "source_evidence": [
            {
                "evidence_id": "EVID-REGISTER",
                "neutral_source_document_id": "SRC-REGISTER",
                "source_role": "MATERIAL_REQUIREMENT",
                "chronology_status": "PRE_DESIGN",
                "generator_input_eligible": True,
                "contains_completed_reference_content": False,
            }
        ],
        "source_facts": [
            {"fact_id": "F1", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R1:C1", "field_type": "model", "fact_type": "model", "component_key": "ROW-1", "value": secret_model, "normalized_value": secret_model, "raw_value": secret_model, "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "F2", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R1:C2", "field_type": "manufacturer", "fact_type": "manufacturer", "component_key": "ROW-1", "value": secret_maker, "normalized_value": secret_maker, "raw_value": secret_maker, "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "F3", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R1:C3", "field_type": "required_qty", "fact_type": "required_qty", "component_key": "ROW-1", "value": 2, "normalized_value": 2, "raw_value": "2", "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "F4", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R2:C1", "field_type": "model", "component_key": "ROW-2", "normalized_value": "REGISTER-SECOND-MODEL", "raw_value": "REGISTER-SECOND-MODEL", "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "F5", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R2:C2", "field_type": "family", "component_key": "ROW-2", "normalized_value": "SYNTH_FAMILY", "raw_value": "SYNTH_FAMILY", "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "F6", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R2:C3", "field_type": "required_qty", "component_key": "ROW-2", "normalized_value": 1, "raw_value": "1", "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
        ],
        "source_line_accounting": [
            {"source_line_id": "ROW-1", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "row_index": 1, "status": "REPRESENTED", "fact_count": 3},
            {"source_line_id": "ROW-2", "evidence_id": "EVID-REGISTER", "neutral_source_document_id": "SRC-REGISTER", "row_index": 2, "status": "REPRESENTED", "fact_count": 3},
        ],
        "quantity_stage_counts": {"required_qty": 2, "ordered_qty": 0, "received_qty": 0, "allocated_qty": 0, "installed_qty": 0},
        "validation": {
            "status": "PASS",
            "evidence_count": 1,
            "source_fact_count": 6,
            "source_line_count": 2,
            "silently_discarded_authorized_source_lines": 0,
            "quantity_stage_overwrite_violations": 0,
            "completed_reference_facts": 0,
            "private_content_transmission_count": 0,
        },
    }
    source_model_path = work / "source_fact_model.json"
    write_json(source_model_path, source_model)
    result = subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--output-dir",
            str(out),
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    assert_true(secret_model not in result.stdout and secret_maker not in result.stdout, "component register values must not be printed to stdout")
    assert_true(secret_model not in result.stderr and secret_maker not in result.stderr, "component register values must not be printed to stderr")
    errors = validate_file(out / "component_register.json", ROOT / "schemas/component_register.schema.json")
    assert_true(not errors, f"component register schema errors: {errors}")
    register = read_json(out / "component_register.json")
    for component_type in register["component_types"]:
        errors = validate(component_type, read_json(ROOT / "schemas/component_type.schema.json"))
        assert_true(not errors, f"component type schema errors: {errors}")
    for component_instance in register["component_instances"]:
        errors = validate(component_instance, read_json(ROOT / "schemas/component_instance.schema.json"))
        assert_true(not errors, f"component instance schema errors: {errors}")
    validation = read_json(out / "component_register_validation.json")
    assert_true(validation["status"] == "PASS", "component register validation must pass")
    assert_true(validation["component_instance_count"] == 2, "two source rows should produce two component instances")
    assert_true(validation["unregistered_allowed_component_key_count"] == 0, "every source fact component key must be registered")
    assert_true(validation["quantity_stage_instance_counts"]["required_qty"] == 2, "required quantities must remain in required_qty")
    assert_true(validation["private_content_transmission_count"] == 0, "private transmission count must stay zero")


def test_sheetmetal_v1_panel_assignment_graph_from_private_models() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_panel_assignment_graph_test"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)
    register_out = work / "register"
    graph_out = work / "graph"
    secret_model = "GRAPH-SECRET-MODEL"
    source_model = {
        "schema_version": "sheetmetal-v1.source_fact_model.v1",
        "project_id": "SYNTH-SMV1-GRAPH",
        "source_mode": "SOURCE_MODE_A_INVENTORY_ONLY",
        "source_evidence": [
            {
                "evidence_id": "EVID-GRAPH",
                "neutral_source_document_id": "SRC-GRAPH",
                "source_role": "PANEL_ALLOCATION_SOURCE",
                "chronology_status": "PRE_DESIGN",
                "generator_input_eligible": True,
                "contains_completed_reference_content": False,
            }
        ],
        "source_facts": [
            {"fact_id": "GF1", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "source_role": "PANEL_ALLOCATION_SOURCE", "source_location_id": "ART:R1:C1", "field_type": "model", "component_key": "ROW-1", "normalized_value": secret_model, "raw_value": secret_model, "authority_class": "SECONDARY_OR_CONTEXT", "confidence": 0.6, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "GF2", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "source_role": "PANEL_ALLOCATION_SOURCE", "source_location_id": "ART:R1:C2", "field_type": "family", "component_key": "ROW-1", "normalized_value": "GRAPH_FAMILY", "raw_value": "GRAPH_FAMILY", "authority_class": "SOURCE_CONTEXT", "confidence": 0.6, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "GF3", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "source_role": "PANEL_ALLOCATION_SOURCE", "source_location_id": "ART:R1:C3", "field_type": "panel_assignment", "component_key": "ROW-1", "normalized_value": "PANEL-A", "raw_value": "PANEL-A", "authority_class": "PRIMARY", "confidence": 0.9, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "GF4", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "source_role": "PANEL_ALLOCATION_SOURCE", "source_location_id": "ART:R2:C1", "field_type": "model", "component_key": "ROW-2", "normalized_value": "GRAPH-SECOND-MODEL", "raw_value": "GRAPH-SECOND-MODEL", "authority_class": "SECONDARY_OR_CONTEXT", "confidence": 0.6, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
        ],
        "source_line_accounting": [
            {"source_line_id": "ROW-1", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "row_index": 1, "status": "REPRESENTED", "fact_count": 3},
            {"source_line_id": "ROW-2", "evidence_id": "EVID-GRAPH", "neutral_source_document_id": "SRC-GRAPH", "row_index": 2, "status": "REPRESENTED", "fact_count": 1},
        ],
        "quantity_stage_counts": {"required_qty": 0, "ordered_qty": 0, "received_qty": 0, "allocated_qty": 0, "installed_qty": 0},
        "validation": {
            "status": "PASS",
            "evidence_count": 1,
            "source_fact_count": 4,
            "source_line_count": 2,
            "silently_discarded_authorized_source_lines": 0,
            "quantity_stage_overwrite_violations": 0,
            "completed_reference_facts": 0,
            "private_content_transmission_count": 0,
        },
    }
    source_model_path = work / "source_fact_model.json"
    write_json(source_model_path, source_model)
    subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--output-dir",
            str(register_out),
            "--quiet",
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    result = subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--component-register",
            str(register_out / "component_register.json"),
            "--output-dir",
            str(graph_out),
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    assert_true(secret_model not in result.stdout, "panel graph summary must not print source values")
    assert_true(secret_model not in result.stderr, "panel graph errors must not print source values")
    errors = validate_file(graph_out / "panel_assignment.json", ROOT / "schemas/panel_assignment.schema.json")
    assert_true(not errors, f"panel assignment schema errors: {errors}")
    errors = validate_file(graph_out / "panel_graph.json", ROOT / "schemas/panel_graph.schema.json")
    assert_true(not errors, f"panel graph schema errors: {errors}")
    validation = read_json(graph_out / "panel_graph_validation.json")
    assert_true(validation["status"] == "PASS", "panel graph validation must pass")
    assert_true(validation["assignment_count"] == 1, "one explicit panel assignment should be accepted")
    assert_true(validation["unresolved_component_count"] == 1, "one component should remain unresolved")
    assert_true(validation["dangling_edge_count"] == 0, "graph edges must resolve to known nodes")
    assert_true(validation["edge_type_counts"]["ASSIGNED_TO_PANEL"] == 1, "panel assignment edge should be present")
    assert_true(validation["edge_type_counts"]["INSTANCE_OF"] == 2, "instance/type edges should be present")
    assert_true(validation["edge_type_counts"]["REQUIRED_BY"] == 2, "required-by edges should be present")
    assert_true(validation["inventory_only_unverified_function_edges"] == 1, "functional graph should remain unverified in inventory-only mode")
    assert_true(validation["private_content_transmission_count"] == 0, "private transmission count must stay zero")


def test_sheetmetal_v1_accessory_cutout_reconciliation_from_private_models() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_accessory_cutout_test"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)
    register_out = work / "register"
    graph_out = work / "graph"
    accessory_out = work / "accessory"
    secret_model = "ACCESSORY-SECRET-MODEL"
    source_model = {
        "schema_version": "sheetmetal-v1.source_fact_model.v1",
        "project_id": "SYNTH-SMV1-ACCESSORY",
        "source_mode": "SOURCE_MODE_A_INVENTORY_ONLY",
        "source_evidence": [
            {
                "evidence_id": "EVID-ACCESSORY",
                "neutral_source_document_id": "SRC-ACCESSORY",
                "source_role": "MATERIAL_REQUIREMENT",
                "chronology_status": "PRE_DESIGN",
                "generator_input_eligible": True,
                "contains_completed_reference_content": False,
            }
        ],
        "source_facts": [
            {"fact_id": "AF1", "evidence_id": "EVID-ACCESSORY", "neutral_source_document_id": "SRC-ACCESSORY", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R1:C1", "field_type": "model", "component_key": "ROW-ACCESSORY-1", "normalized_value": secret_model, "raw_value": secret_model, "authority_class": "SECONDARY_OR_CONTEXT", "confidence": 0.6, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
            {"fact_id": "AF2", "evidence_id": "EVID-ACCESSORY", "neutral_source_document_id": "SRC-ACCESSORY", "source_role": "MATERIAL_REQUIREMENT", "source_location_id": "ART:R1:C2", "field_type": "family", "component_key": "ROW-ACCESSORY-1", "normalized_value": "ACCESSORY_SOURCE_FAMILY", "raw_value": "ACCESSORY_SOURCE_FAMILY", "authority_class": "SOURCE_CONTEXT", "confidence": 0.6, "chronology_status": "PRE_DESIGN", "conflict_status": "NONE", "status": "EXPLICIT_SOURCE"},
        ],
        "source_line_accounting": [
            {"source_line_id": "ROW-ACCESSORY-1", "evidence_id": "EVID-ACCESSORY", "neutral_source_document_id": "SRC-ACCESSORY", "row_index": 1, "status": "REPRESENTED", "fact_count": 2},
        ],
        "accessory_rules": [
            {
                "rule_id": "AR-SYNTH-1",
                "version": "v1",
                "applicable_component_family": "ACCESSORY_SOURCE_FAMILY",
                "condition": "synthetic current-project rule",
                "generated_requirement": {"family": "SYNTH_ACCESSORY", "model": "SYNTH-BRACKET"},
                "generated_cutout": {"shape": "rectangle", "width_mm": 20, "height_mm": 10, "units": "mm"},
                "evidence_basis": "synthetic unit test rule",
                "confidence": 0.8,
                "priority": 1,
                "conflict_behavior": "do_not_duplicate_explicit_accessory",
            }
        ],
        "quantity_stage_counts": {"required_qty": 0, "ordered_qty": 0, "received_qty": 0, "allocated_qty": 0, "installed_qty": 0},
        "validation": {
            "status": "PASS",
            "evidence_count": 1,
            "source_fact_count": 2,
            "source_line_count": 1,
            "silently_discarded_authorized_source_lines": 0,
            "quantity_stage_overwrite_violations": 0,
            "completed_reference_facts": 0,
            "private_content_transmission_count": 0,
        },
    }
    for rule in source_model["accessory_rules"]:
        errors = validate(rule, read_json(ROOT / "schemas/accessory_rule.schema.json"))
        assert_true(not errors, f"accessory rule schema errors: {errors}")
    source_model_path = work / "source_fact_model.json"
    write_json(source_model_path, source_model)
    subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--output-dir",
            str(register_out),
            "--quiet",
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--component-register",
            str(register_out / "component_register.json"),
            "--output-dir",
            str(graph_out),
            "--quiet",
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    result = subprocess.run(
        [
            PY,
            "scripts/sheetmetal_v1.py",
            "--source-fact-model",
            str(source_model_path),
            "--component-register",
            str(register_out / "component_register.json"),
            "--panel-graph",
            str(graph_out / "panel_graph.json"),
            "--output-dir",
            str(accessory_out),
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=True,
    )
    assert_true(secret_model not in result.stdout, "accessory summary must not print source values")
    assert_true(secret_model not in result.stderr, "accessory errors must not print source values")
    accessories = read_json(accessory_out / "accessory_requirements.json")
    for component_instance in accessories["generated_component_instances"]:
        errors = validate(component_instance, read_json(ROOT / "schemas/component_instance.schema.json"))
        assert_true(not errors, f"generated accessory component instance schema errors: {errors}")
    validation = read_json(accessory_out / "accessory_cutout_validation.json")
    assert_true(validation["status"] == "PASS", "accessory/cutout validation must pass")
    assert_true(validation["requirement_count"] == 1, "one accessory requirement should be generated")
    assert_true(validation["generated_component_instance_count"] == 1, "one accessory component should be generated")
    assert_true(validation["cutout_count"] == 1, "one cutout should be generated")
    assert_true(validation["duplicate_accessory_count"] == 0, "no duplicate accessory should be reported")
    assert_true(validation["missing_requirement_source_count"] == 0, "requirement source must resolve in graph")
    assert_true(validation["missing_cutout_source_count"] == 0, "cutout source must resolve in graph")
    assert_true(validation["private_content_transmission_count"] == 0, "private transmission count must stay zero")


def test_sheetmetal_v1_topology_sizing_placement_calibration_private_models() -> None:
    work = ROOT / "tmp" / "sheetmetal_v1_topology_calibration_test"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)
    register_out = work / "register"
    graph_out = work / "graph"
    accessory_out = work / "accessory"
    topo_out = work / "topology"
    topo_rerun = work / "topology_rerun"
    secret_model = "TOPOLOGY-SECRET-MODEL"

    def fact(fid: str, evid: str, component_key: str, field_type: str, value, role: str = "MATERIAL_REQUIREMENT") -> dict:
        return {
            "fact_id": fid,
            "evidence_id": evid,
            "neutral_source_document_id": f"SRC-{evid}",
            "source_role": role,
            "source_location_id": f"ART:{fid}",
            "field_type": field_type,
            "component_key": component_key,
            "normalized_value": value,
            "raw_value": value,
            "authority_class": "PRIMARY" if field_type == "panel_assignment" and role == "PANEL_ALLOCATION_SOURCE" else "SOURCE_CONTEXT",
            "confidence": 0.8,
            "chronology_status": "PRE_DESIGN" if role != "POST_DESIGN_ALLOCATION_LABEL" else "POST_DESIGN",
            "conflict_status": "NONE",
            "status": "EXPLICIT_SOURCE",
        }

    source_model = {
        "schema_version": "sheetmetal-v1.source_fact_model.v1",
        "project_id": "SYNTH-SMV1-TOPO",
        "source_mode": "SOURCE_MODE_A_INVENTORY_ONLY",
        "source_evidence": [
            {
                "evidence_id": "EVID-MAT",
                "neutral_source_document_id": "SRC-MAT",
                "source_role": "MATERIAL_REQUIREMENT",
                "chronology_status": "PRE_DESIGN",
                "generator_input_eligible": True,
                "contains_completed_reference_content": False,
            },
            {
                "evidence_id": "EVID-PANEL",
                "neutral_source_document_id": "SRC-PANEL",
                "source_role": "PANEL_ALLOCATION_SOURCE",
                "chronology_status": "PRE_DESIGN",
                "generator_input_eligible": True,
                "contains_completed_reference_content": False,
            },
            {
                "evidence_id": "EVID-POST",
                "neutral_source_document_id": "SRC-POST",
                "source_role": "POST_DESIGN_ALLOCATION_LABEL",
                "chronology_status": "POST_DESIGN",
                "generator_input_eligible": False,
                "contains_completed_reference_content": False,
            },
            {
                "evidence_id": "EVID-REF",
                "neutral_source_document_id": "SRC-REF",
                "source_role": "COMPLETED_SHEETMETAL_REFERENCE",
                "chronology_status": "AFTER_COMPLETION",
                "generator_input_eligible": False,
                "contains_completed_reference_content": True,
            },
        ],
        "source_facts": [],
        "source_line_accounting": [],
        "panel_definitions": [
            {"panel_id": "PANEL-A", "width_mm": 100, "height_mm": 100, "depth_mm": 30, "mounting_surfaces": [{"surface_id": "SURF-A"}], "evidence_ids": ["EVID-PANEL"], "status": "EXPLICIT_SOURCE"},
            {"panel_id": "PANEL-B", "width_mm": 120, "height_mm": 100, "depth_mm": 30, "mounting_surfaces": [{"surface_id": "SURF-B"}], "evidence_ids": ["EVID-PANEL"], "status": "EXPLICIT_SOURCE"},
            {"panel_id": "PANEL-C", "mounting_surfaces": [], "evidence_ids": [], "status": "HUMAN_REVIEW_REQUIRED"},
        ],
        "placement_requests": [
            {"component_key": "ROW-OK", "panel_id": "PANEL-A", "mounting_surface_id": "SURF-A", "x_mm": 10, "y_mm": 10, "width_mm": 20, "height_mm": 10, "orientation": "NORMAL", "minimum_edge_clearance_mm": 5, "choice_id": "CHOICE-OK"},
            {"component_key": "ROW-OVERLAP", "panel_id": "PANEL-A", "mounting_surface_id": "SURF-A", "x_mm": 15, "y_mm": 12, "width_mm": 20, "height_mm": 10, "orientation": "NORMAL", "soft_objective_override_attempted": True},
            {"component_key": "ROW-CONTAIN", "panel_id": "PANEL-A", "mounting_surface_id": "SURF-A", "x_mm": 95, "y_mm": 95, "width_mm": 20, "height_mm": 10, "orientation": "NORMAL", "soft_objective_override_attempted": True},
            {"component_key": "ROW-CLEAR", "panel_id": "PANEL-A", "mounting_surface_id": "SURF-A", "x_mm": 2, "y_mm": 2, "width_mm": 20, "height_mm": 10, "orientation": "NORMAL", "minimum_edge_clearance_mm": 5, "soft_objective_override_attempted": True},
        ],
        "accessory_rules": [],
        "quantity_stage_counts": {"required_qty": 0, "ordered_qty": 0, "received_qty": 0, "allocated_qty": 0, "installed_qty": 0},
        "validation": {"status": "PASS", "evidence_count": 4, "source_fact_count": 0, "source_line_count": 0, "silently_discarded_authorized_source_lines": 0, "quantity_stage_overwrite_violations": 0, "completed_reference_facts": 0, "private_content_transmission_count": 0},
    }
    rows = [
        ("ROW-OK", secret_model, {"width_mm": 20, "height_mm": 10, "depth_mm": 5}, "PANEL-A"),
        ("ROW-MISSING-GEOM", "MISSING-GEOM", None, "PANEL-A"),
        ("ROW-UNASSIGNED", "UNASSIGNED", {"width_mm": 10, "height_mm": 10, "depth_mm": 5}, None),
        ("ROW-OVERLAP", "OVERLAP", {"width_mm": 20, "height_mm": 10, "depth_mm": 5}, "PANEL-A"),
        ("ROW-CONTAIN", "CONTAIN", {"width_mm": 20, "height_mm": 10, "depth_mm": 5}, "PANEL-A"),
        ("ROW-CLEAR", "CLEAR", {"width_mm": 20, "height_mm": 10, "depth_mm": 5}, "PANEL-A"),
        ("ROW-POST", "POST", {"width_mm": 10, "height_mm": 10, "depth_mm": 5}, None),
    ]
    for index, (component_key, model, geometry, panel_id) in enumerate(rows, start=1):
        source_model["source_facts"].append(fact(f"TF{index}A", "EVID-MAT", component_key, "family", "TOPO_FAMILY"))
        source_model["source_facts"].append(fact(f"TF{index}B", "EVID-MAT", component_key, "model", model))
        if geometry:
            source_model["source_facts"].append(fact(f"TF{index}C", "EVID-MAT", component_key, "component_geometry", geometry))
        if panel_id:
            source_model["source_facts"].append(fact(f"TF{index}D", "EVID-PANEL", component_key, "panel_assignment", panel_id, "PANEL_ALLOCATION_SOURCE"))
        source_model["source_line_accounting"].append({"source_line_id": component_key, "evidence_id": "EVID-MAT", "neutral_source_document_id": "SRC-MAT", "row_index": index, "status": "REPRESENTED", "fact_count": 2})
    source_model["source_facts"].append(fact("TFPOST", "EVID-POST", "ROW-POST", "panel_assignment", "PANEL-A", "POST_DESIGN_ALLOCATION_LABEL"))
    source_model["source_facts"].append(fact("TFREF", "EVID-REF", "ROW-MISSING-GEOM", "component_geometry", {"width_mm": 99, "height_mm": 99, "depth_mm": 99}, "COMPLETED_SHEETMETAL_REFERENCE"))
    source_model["validation"]["source_fact_count"] = len(source_model["source_facts"])
    source_model["validation"]["source_line_count"] = len(source_model["source_line_accounting"])

    source_model_path = work / "source_fact_model.json"
    write_json(source_model_path, source_model)
    subprocess.run([PY, "scripts/sheetmetal_v1.py", "--source-fact-model", str(source_model_path), "--output-dir", str(register_out), "--quiet"], cwd=ROOT, check=True)
    register = read_json(register_out / "component_register.json")
    instance_by_key = {row["component_key"]: row for row in register["component_instances"]}
    generic_type_id = instance_by_key["ROW-CLEAR"]["component_type_id"]
    for ctype in register["component_types"]:
        if ctype["component_type_id"] == generic_type_id:
            ctype["geometry"]["status"] = "APPROVED_GENERIC_CONSERVATIVE_ENVELOPE"
    write_json(register_out / "component_register.json", register)
    subprocess.run([PY, "scripts/sheetmetal_v1.py", "--source-fact-model", str(source_model_path), "--component-register", str(register_out / "component_register.json"), "--output-dir", str(graph_out), "--quiet"], cwd=ROOT, check=True)
    subprocess.run([PY, "scripts/sheetmetal_v1.py", "--source-fact-model", str(source_model_path), "--component-register", str(register_out / "component_register.json"), "--panel-graph", str(graph_out / "panel_graph.json"), "--output-dir", str(accessory_out), "--quiet"], cwd=ROOT, check=True)

    cmd = [
        PY,
        "scripts/sheetmetal_v1.py",
        "--topology-calibration",
        "--source-fact-model",
        str(source_model_path),
        "--component-register",
        str(register_out / "component_register.json"),
        "--panel-assignment",
        str(graph_out / "panel_assignment.json"),
        "--panel-graph",
        str(graph_out / "panel_graph.json"),
        "--accessory-requirements",
        str(accessory_out / "accessory_requirements.json"),
    ]
    result = subprocess.run([*cmd, "--output-dir", str(topo_out)], cwd=ROOT, text=True, capture_output=True, check=True)
    subprocess.run([*cmd, "--output-dir", str(topo_rerun), "--quiet"], cwd=ROOT, text=True, capture_output=True, check=True)
    assert_true(secret_model not in result.stdout, "topology summary must not print source values")
    assert_true(secret_model not in result.stderr, "topology errors must not print source values")

    assignment = read_json(topo_out / "panel_assignment_recovery.json")
    assert_true(assignment["counts"]["unsupported_assignment_count"] == 0, "unsupported assignments must remain zero")
    assert_true(assignment["rejected_source_assignment_count"] == 1, "post-design assignment must be rejected before recovery")
    assert_true(assignment["counts"]["unassigned"] >= 1, "unassigned components must stay explicit")
    topology = read_json(topo_out / "topology_candidates.json")
    assert_true(topology["candidate_count"] == 3, "multiple topology candidates must be preserved")
    assert_true(topology["multiple_valid_candidates_preserved"], "valid candidates must not be silently collapsed")
    sizing = read_json(topo_out / "sizing_candidates.json")
    assert_true(sizing["geometry_status_counts"]["GEOMETRY_MISSING"] >= 1, "missing geometry must remain missing")
    assert_true(sizing["geometry_status_counts"]["APPROVED_GENERIC_CONSERVATIVE_ENVELOPE"] >= 1, "generic geometry must not become verified")
    assert_true(any(row["rejected_alternatives"] for row in sizing["candidates"]), "unsupported exact cabinet sizes must be blocked")
    placement = read_json(topo_out / "placement_plan.json")
    unplaced = read_json(topo_out / "unplaced_component_register.json")
    constraints = read_json(topo_out / "hard_constraint_model.json")
    validation = read_json(topo_out / "validation_report.json")
    provenance = read_json(topo_out / "provenance_map.json")
    assert_true(placement["placement_count"] == 1, "only hard-constraint-valid placement should be accepted")
    assert_true(unplaced["reason_counts"]["UNASSIGNED_PANEL"] >= 1, "unassigned components must remain unplaced")
    assert_true(unplaced["reason_counts"]["GEOMETRY_MISSING"] >= 1, "missing geometry must remain unplaced")
    assert_true(unplaced["reason_counts"]["NO_VALID_PLACEMENT"] >= 1, "invalid placements must stay unplaced")
    flattened = {item for row in constraints["rejected_placements"] for item in row["unsatisfied_constraints"]}
    assert_true({"NO_PHYSICAL_OVERLAP", "CONTAINMENT", "EDGE_CLEARANCE"} <= flattened, "overlap, containment, and clearance must be detected")
    assert_true(any(row["soft_objective_override_rejected"] for row in constraints["rejected_placements"]), "hard constraints must override soft objectives")
    assert_true(validation["unsupported_critical_dimensions"] == 0, "unsupported critical dimensions must remain zero")
    assert_true(validation["completed_reference_leakage"] == 0, "completed-reference leakage must remain zero")
    assert_true(validation["post_design_leakage"] == 0, "post-design leakage must remain zero")
    assert_true(validation["customer_drawing_generation_count"] == 0, "topology calibration must not generate drawings")
    assert_true(provenance["coverage_status"] == "PASS", "provenance must cover critical facts or safe unresolved statuses")
    for path in topo_out.iterdir():
        assert_true(path.suffix.lower() not in {".pdf", ".dxf", ".dwg"}, "no drawing artifact may be generated")
    assert_true(not (topo_out / "sheetmetal_drawing_model.json").exists(), "drawing model must not be generated in topology calibration mode")
    for name in [
        "panel_assignment_recovery",
        "topology_candidates",
        "sizing_candidates",
        "placement_plan",
        "unplaced_component_register",
        "hard_constraint_model",
        "validation_report",
        "provenance_map",
    ]:
        assert_true(sha256_json(read_json(topo_out / f"{name}.json")) == sha256_json(read_json(topo_rerun / f"{name}.json")), f"{name} must be deterministic")


def test_frozen_workflow_legacy_scope_verifier() -> None:
    import verify_frozen_workflow as verifier

    manifest_path = ROOT / "evals" / "baseline-024" / "frozen_workflow_manifest.json"
    manifest_before = manifest_path.read_bytes()
    attestation = verifier.build_legacy_attestation(worktree_name="legacy-baseline-024-test-attest")
    assert_true(attestation["status"] == "PASS", f"legacy attestation should pass: {attestation}")
    assert_true(attestation["verified_file_count"] == 15, "legacy attestation must verify every frozen file")
    assert_true(
        any(
            row["key"] == "agents_md"
            and row["expected_sha256"] == "6298E5EEB469DEBA38CD3F83D5697C402815B7E0C94F61E7F355D6D704D8E433"
            and row["status"] == "PASS"
            for row in attestation["hashes"]
        ),
        "historical AGENTS.md hash must reproduce at the legacy anchor",
    )
    current_agents_hash = verifier.sha256_file(ROOT / "AGENTS.md")
    assert_true(
        current_agents_hash != "6298E5EEB469DEBA38CD3F83D5697C402815B7E0C94F61E7F355D6D704D8E433",
        "test assumes current sheetmetal-v1 AGENTS.md differs from legacy hash",
    )
    result = verifier.verify_legacy_scope(attestation=attestation, worktree_name="legacy-baseline-024-test-verify")
    assert_true(result["status"] == "PASS", f"legacy scoped verifier must not compare current AGENTS.md: {result}")
    assert_true(not (ROOT / "tmp" / "frozen_workflow_verification").exists(), "temporary legacy worktree root must be removed")
    assert_true(manifest_path.read_bytes() == manifest_before, "old legacy manifest must remain byte-identical")


def test_frozen_workflow_fail_closed_regressions() -> None:
    import verify_frozen_workflow as verifier

    fake_root = ROOT / "tmp" / "frozen_workflow_unit" / "hash_root"
    if fake_root.exists():
        shutil.rmtree(fake_root)
    fake_root.mkdir(parents=True)
    sample = fake_root / "sample.txt"
    sample.write_text("accepted\n", encoding="utf-8", newline="\n")
    expected = verifier.sha256_file(sample)
    checks = verifier.hash_checks_in_directory(fake_root, {"sample": expected}, {"sample": "sample.txt"})
    assert_true(checks[0]["status"] == "PASS", "matching historical file should pass")
    sample.write_text("altered\n", encoding="utf-8", newline="\n")
    checks = verifier.hash_checks_in_directory(fake_root, {"sample": expected}, {"sample": "sample.txt"})
    assert_true(checks[0]["status"] == "FAIL", "altered historical file must fail")

    attestation = verifier.build_legacy_attestation(worktree_name="legacy-baseline-024-test-failclosed")
    mutated_hash = json.loads(json.dumps(attestation))
    mutated_hash["manifest_sha256"] = "0" * 64
    result = verifier.verify_legacy_scope(attestation=mutated_hash, worktree_name="legacy-baseline-024-test-mutatedhash")
    assert_true(result["status"] == "FAIL" and "MANIFEST_HASH_MISMATCH" in result["failures"], "legacy manifest hash mutation must fail")

    missing_anchor = json.loads(json.dumps(attestation))
    missing_anchor["resolved_historical_anchor_commit"] = "0" * 40
    result = verifier.verify_legacy_scope(attestation=missing_anchor, worktree_name="legacy-baseline-024-test-missinganchor")
    assert_true(result["status"] == "FAIL" and "MISSING_ANCHOR_COMMIT" in result["failures"], "missing historical anchor must fail closed")

    ambiguous = verifier.resolve_unique_legacy_anchor([
        {"commit": "a" * 40, "status": "PASS"},
        {"commit": "b" * 40, "status": "PASS"},
    ])
    assert_true(ambiguous["status"] == "FAIL" and ambiguous["error"] == "AMBIGUOUS_LEGACY_ANCHOR", "ambiguous anchor must fail closed")


def test_frozen_workflow_active_scope_verifier() -> None:
    import verify_frozen_workflow as verifier

    if verifier.ACTIVE_MANIFEST.exists():
        manifest = read_json(verifier.ACTIVE_MANIFEST)
        result = verifier.verify_active_scope()
        assert_true(result["status"] == "PASS", f"active sheetmetal-v1 manifest must pass: {result}")
    else:
        manifest = {
            "manifest_version": "frozen-workflow-v2",
            "scope": "SHEETMETAL_V1_ACTIVE",
            "active_goal": "SHEETMETAL_FIRST_MODULAR_PANEL_MODEL_V1",
            "anchor_commit": verifier.current_head(),
            "hash_algorithm": "SHA-256",
            "supersedes_historical_manifest": False,
            "files": [{"path": "AGENTS.md", "sha256": verifier.sha256_file(ROOT / "AGENTS.md")}],
        }
        result = verifier.verify_active_manifest_data(manifest, manifest_path=ROOT / "tmp" / "active_manifest_fixture.json")
        assert_true(result["status"] == "PASS", f"active manifest fixture should pass: {result}")

    mutated_agents = json.loads(json.dumps(manifest))
    for row in mutated_agents["files"]:
        if row["path"] == "AGENTS.md":
            row["sha256"] = "0" * 64
    result = verifier.verify_active_manifest_data(mutated_agents, manifest_path=ROOT / "tmp" / "active_manifest_mutated.json")
    assert_true(result["status"] == "FAIL" and any("AGENTS.md" in failure for failure in result["failures"]), "active AGENTS.md hash mutation must fail")

    cross_scope = read_json(ROOT / "evals" / "baseline-024" / "frozen_workflow_manifest.json")
    result = verifier.verify_active_manifest_data(cross_scope, manifest_path=ROOT / "evals" / "baseline-024" / "frozen_workflow_manifest.json")
    assert_true(result["status"] == "FAIL" and "SCOPE_MISMATCH" in result["failures"], "cross-scope manifest use must fail")

    missing_anchor = json.loads(json.dumps(manifest))
    missing_anchor["anchor_commit"] = "0" * 40
    result = verifier.verify_active_manifest_data(missing_anchor, manifest_path=ROOT / "tmp" / "active_manifest_missing_anchor.json")
    assert_true(result["status"] == "FAIL" and "MISSING_ANCHOR_COMMIT" in result["failures"], "active missing anchor must fail closed")

    dynamic = json.loads(json.dumps(manifest))
    dynamic["files"].append({"path": "docs/01_CURRENT_STATE.md", "sha256": verifier.sha256_file(ROOT / "docs" / "01_CURRENT_STATE.md")})
    result = verifier.verify_active_manifest_data(dynamic, manifest_path=ROOT / "tmp" / "active_manifest_dynamic.json")
    assert_true(result["status"] == "FAIL" and any(failure.startswith("DYNAMIC_FILE_INCLUDED") for failure in result["failures"]), "dynamic state files must stay out of active workflow freeze")


def test_frozen_workflow_topology_scope_verifier() -> None:
    import verify_frozen_workflow as verifier

    if verifier.TOPOLOGY_MANIFEST.exists():
        result = verifier.verify_topology_scope()
        assert_true(result["status"] == "PASS", f"topology stage manifest must pass: {result}")
        manifest = read_json(verifier.TOPOLOGY_MANIFEST)
    else:
        manifest = {
            "manifest_version": "frozen-workflow-v2",
            "scope": verifier.TOPOLOGY_SCOPE,
            "active_goal": "SHEETMETAL_FIRST_MODULAR_PANEL_MODEL_V1",
            "anchor_commit": verifier.current_head(),
            "hash_algorithm": "SHA-256",
            "supersedes_historical_manifest": False,
            "files": [{"path": "scripts/sheetmetal_v1.py", "sha256": verifier.sha256_file(ROOT / "scripts" / "sheetmetal_v1.py")}],
        }
        result = verifier.verify_active_manifest_data(manifest, manifest_path=ROOT / "tmp" / "topology_manifest_fixture.json", expected_scope=verifier.TOPOLOGY_SCOPE)
        assert_true(result["status"] == "PASS", f"topology manifest fixture should pass: {result}")

    wrong_scope = json.loads(json.dumps(manifest))
    wrong_scope["scope"] = "SHEETMETAL_V1_ACTIVE"
    result = verifier.verify_active_manifest_data(wrong_scope, manifest_path=ROOT / "tmp" / "topology_manifest_wrong_scope.json", expected_scope=verifier.TOPOLOGY_SCOPE)
    assert_true(result["status"] == "FAIL" and "SCOPE_MISMATCH" in result["failures"], "topology scope mismatch must fail")


def main() -> None:
    tests = [
        test_json_schemas_parse,
        test_declared_json_artifacts_validate,
        test_utf8_fixture_labels,
        test_forbidden_classification,
        test_grading_weights,
        test_evaluator_sensitivity_monotonicity,
        test_synthetic_render_grade,
        test_positive_bundle_and_contamination_scan,
        test_source_guard_fail_closed_decisions,
        test_source_approval_and_bundle_fail_closed,
        test_signed_authority_decision_validator_fail_closed,
        test_signed_authority_decision_intake_routing,
        test_signed_authority_decision_draft_scaffold_fail_closed,
        test_signed_authority_decision_submission_package,
        test_bundle_verifier_rejects_leaks_and_traversal,
        test_reference_detection_v3_boundary_and_page_level_sets,
        test_reference_detection_v3_duplicates_identity_and_missing_types,
        test_reference_detector_v3_known_positive_recall_gate,
        test_windows_media_ocr_probe_minimized,
        test_reference_detection_v4_multisignal_and_conflicts,
        test_reference_detection_v4_image_only_ocr_and_failure_modes,
        test_reference_detection_v4_private_ocr_disabled_layout_prior,
        test_qualification_recovery_controller_state,
        test_sheetmetal_v1_modular_foundation,
        test_sheetmetal_v1_private_workspace_boundary,
        test_sheetmetal_v1_source_fact_extractor,
        test_sheetmetal_v1_component_register_from_source_facts,
        test_sheetmetal_v1_panel_assignment_graph_from_private_models,
        test_sheetmetal_v1_accessory_cutout_reconciliation_from_private_models,
        test_sheetmetal_v1_topology_sizing_placement_calibration_private_models,
        test_frozen_workflow_legacy_scope_verifier,
        test_frozen_workflow_fail_closed_regressions,
        test_frozen_workflow_active_scope_verifier,
        test_frozen_workflow_topology_scope_verifier,
    ]
    failures = []
    for test in tests:
        try:
            test()
            print(f"PASS {test.__name__}")
        except Exception as exc:
            failures.append((test.__name__, str(exc)))
            print(f"FAIL {test.__name__}: {exc}")
    if failures:
        print(json.dumps({"status": "FAIL", "failures": failures}, ensure_ascii=False, indent=2))
        raise SystemExit(1)
    print(json.dumps({"status": "PASS", "tests": [t.__name__ for t in tests]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
