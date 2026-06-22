from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from harness_lib import detect_forbidden_text, read_json, sha256_file, sha256_json, write_json


POLICY_VERSION = "source_guard_policy_v1"
SOURCE_ROOT_ID = "SRC-ALL-PROJECTS"
SOURCE_ROOT_PATH = Path(r"C:\Users\alex1\OneDrive\Desktop\All Projects")
PROJECT_ID_RE = re.compile(r"(?<!\d)(1[0-9]{2}[0-9]{4})(?!\d)")
DATE_RE = re.compile(r"(1[0-9]{2}[./-]\d{1,2}[./-]\d{1,2}|\d{4}[./-]\d{1,2}[./-]\d{1,2})")
IDENTITY_TERMS = ["合約", "材料清單", "生管文件", "生管課用圖", "生管用圖", "沖孔", "鈑金", "電機施工圖"]
SAFE_HUMAN_DECISIONS = {"HUMAN_APPROVED", "HUMAN_DENIED", "NEEDS_MORE_EVIDENCE"}


def csv_rows(path: Path) -> list[dict[str, str]]:
    csv.field_size_limit(1024 * 1024 * 128)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def csv_write(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(text.split())


def source_path_from_row(file_row: dict[str, str]) -> Path:
    absolute = file_row.get("absolute_path", "")
    if absolute:
        return Path(absolute)
    return SOURCE_ROOT_PATH / file_row["relative_path"]


def is_inside_source_root(path: Path) -> bool:
    try:
        path.resolve().relative_to(SOURCE_ROOT_PATH.resolve())
        return True
    except Exception:
        return False


def current_hash_matches(file_row: dict[str, str]) -> bool:
    path = source_path_from_row(file_row)
    return path.exists() and is_inside_source_root(path) and sha256_file(path).upper() == file_row.get("sha256", "").upper()


def openpyxl_version() -> str:
    try:
        import openpyxl

        return openpyxl.__version__
    except Exception:
        return "unavailable"


def extract_sheet_refs(formula: str) -> list[str]:
    refs = []
    for match in re.finditer(r"(?:'([^']+)'|([A-Za-z0-9_\-\u4e00-\u9fff ]+))!", formula):
        refs.append((match.group(1) or match.group(2) or "").strip())
    return sorted(set(refs))


def worksheet_fingerprint(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=False, data_only=False, keep_links=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    cells: list[dict[str, str]] = []
    formulas: list[dict[str, Any]] = []
    identity_cells: list[dict[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            normalized = normalize_value(cell.value)
            if not normalized:
                continue
            entry = {"coordinate": cell.coordinate, "type": type(cell.value).__name__, "value": normalized}
            cells.append(entry)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({"coordinate": cell.coordinate, "formula": normalized, "sheet_refs": extract_sheet_refs(normalized)})
            if PROJECT_ID_RE.search(normalized) or DATE_RE.search(normalized) or any(term in normalized for term in IDENTITY_TERMS):
                identity_cells.append(entry)

    defined_names = []
    try:
        for name, defn in wb.defined_names.items():
            text = str(defn.attr_text)
            if sheet_name in text:
                defined_names.append({"name": name, "refers_to": text})
    except Exception:
        defined_names = []

    payload = {
        "sheet_name": ws.title,
        "sheet_visibility": ws.sheet_state,
        "workbook_sheet_count": len(wb.sheetnames),
        "used_range_dimensions": ws.calculate_dimension(),
        "normalized_visible_cell_content": cells,
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
        "formula_locations": formulas,
        "defined_names_touching_sheet": defined_names,
        "identity_like_cells": identity_cells,
    }
    return {
        "fingerprint": sha256_json(payload),
        "payload": payload,
        "formula_count": len(formulas),
        "formula_sheet_refs": sorted({ref for item in formulas for ref in item["sheet_refs"]}),
        "defined_name_count": len(defined_names),
        "external_link_count": len(getattr(wb, "_external_links", []) or []),
        "sheet_count": len(wb.sheetnames),
    }


def family_id(project_name: str) -> str:
    if not project_name:
        return "UNKNOWN"
    clean = re.sub(r"^1[0-9]{6}", "", project_name).strip()
    if "-" in clean:
        return clean.split("-", 1)[0].strip() or "UNKNOWN"
    return clean[:4] or "UNKNOWN"


def detected_ids(*texts: str) -> list[str]:
    ids: set[str] = set()
    for text in texts:
        ids.update(PROJECT_ID_RE.findall(text or ""))
    return sorted(ids)


def decision_id(file_id: str, sheet_id: str | None = None) -> str:
    base = f"{file_id}:{sheet_id or 'FILE'}"
    return "SGD-" + hashlib.sha1(base.encode("utf-8")).hexdigest()[:14].upper()


def decide_source_item(file_row: dict[str, str], sheet_row: dict[str, str] | None = None, compute_fingerprint: bool = True) -> dict[str, Any]:
    path = source_path_from_row(file_row)
    ext = (file_row.get("extension") or path.suffix).lower()
    text_scope = " ".join([
        file_row.get("relative_path", ""),
        file_row.get("file_name", ""),
        file_row.get("primary_role", ""),
        sheet_row.get("sheet_name", "") if sheet_row else "",
        sheet_row.get("stale_evidence", "") if sheet_row else "",
    ])
    reasons: list[str] = []
    proposed = "NEEDS_HUMAN_REVIEW"
    parser_name = "none"
    parser_version = "none"
    parse_status = file_row.get("parse_status", "NOT_PARSED")
    worksheet_name = sheet_row.get("sheet_name", "") if sheet_row else ""
    worksheet_visibility = sheet_row.get("visibility", "") if sheet_row else ""
    worksheet_fp = ""
    sheet_count: str | int | None = ""
    formula_status = "NONE"
    external_status = "NONE"
    named_range_status = "NONE"
    revision_status = "CURRENT_OR_UNKNOWN"
    duplicate = file_row.get("duplicate_of") or ""

    if not path.exists() or not is_inside_source_root(path):
        reasons.append("OUTSIDE_APPROVED_ROOT_OR_MISSING")
        proposed = "AUTO_DENIED"
    elif sha256_file(path).upper() != file_row.get("sha256", "").upper():
        reasons.append("SOURCE_HASH_CHANGED")
        proposed = "AUTO_DENIED"

    forbidden_hits = detect_forbidden_text(text_scope)
    if forbidden_hits or file_row.get("generator_eligibility") == "FORBIDDEN" or (sheet_row and sheet_row.get("generator_eligibility") == "FORBIDDEN"):
        reasons.append("FORBIDDEN_LABEL_OR_ROLE")
        proposed = "AUTO_DENIED"
    if "修改" in text_scope or "revision" in text_scope.lower():
        revision_status = "REVISION_REVIEW_REQUIRED"
        reasons.append("REVISION_OR_MODIFICATION_SIGNAL")
        if proposed != "AUTO_DENIED":
            proposed = "QUARANTINED"

    if sheet_row:
        stale = sheet_row.get("stale_template_status", "")
        if stale == "STALE_TEMPLATE_SHEET":
            reasons.append("STALE_TEMPLATE_SHEET")
            proposed = "AUTO_DENIED"
        elif stale == "FORBIDDEN_SOURCE":
            reasons.append("FORBIDDEN_SOURCE")
            proposed = "AUTO_DENIED"
        elif stale in {"INSUFFICIENT_IDENTITY", "HUMAN_REVIEW_REQUIRED"}:
            reasons.append(stale)
        if worksheet_visibility in {"hidden", "veryHidden"}:
            reasons.append("HIDDEN_OR_VERY_HIDDEN_SHEET")
            if proposed != "AUTO_DENIED":
                proposed = "QUARANTINED"
        if sheet_row.get("external_links"):
            external_status = "EXTERNAL_LINK_REVIEW_REQUIRED"
            reasons.append("EXTERNAL_LINK_PRESENT")
            if proposed != "AUTO_DENIED":
                proposed = "QUARANTINED"
        if sheet_row.get("named_ranges"):
            named_range_status = "NAMED_RANGE_REVIEW_REQUIRED"
            reasons.append("NAMED_RANGE_PRESENT")
            if proposed != "AUTO_DENIED":
                proposed = "QUARANTINED"

    if ext == ".xls":
        parser_name = "parser_required"
        parser_version = "none"
        parse_status = "PARSER_REQUIRED"
        reasons.append("LEGACY_XLS_PARSER_REQUIRED")
        if proposed != "AUTO_DENIED":
            proposed = "PARSER_REQUIRED"
    elif ext in {".xlsx", ".xlsm"} and sheet_row and worksheet_name:
        parser_name = "openpyxl"
        parser_version = openpyxl_version()
        if ext == ".xlsm":
            reasons.append("MACRO_ENABLED_WORKBOOK_QUARANTINE")
            if proposed != "AUTO_DENIED":
                proposed = "QUARANTINED"
        if compute_fingerprint and proposed not in {"AUTO_DENIED", "PARSER_REQUIRED"}:
            try:
                fp = worksheet_fingerprint(path, worksheet_name)
                worksheet_fp = fp["fingerprint"]
                sheet_count = fp["sheet_count"]
                parse_status = "PARSED_WITH_FINGERPRINT"
                if fp["external_link_count"]:
                    external_status = "EXTERNAL_LINK_REVIEW_REQUIRED"
                    reasons.append("EXTERNAL_LINK_PRESENT")
                    if proposed != "AUTO_DENIED":
                        proposed = "QUARANTINED"
                if fp["defined_name_count"]:
                    named_range_status = "NAMED_RANGE_REVIEW_REQUIRED"
                    reasons.append("NAMED_RANGE_PRESENT")
                    if proposed != "AUTO_DENIED":
                        proposed = "QUARANTINED"
                denied_refs = [ref for ref in fp["formula_sheet_refs"] if ref and ref != worksheet_name]
                if denied_refs:
                    formula_status = "FORMULA_DEPENDENCY_REVIEW_REQUIRED"
                    reasons.append("FORMULA_REFERENCES_OTHER_SHEET")
                    if proposed != "AUTO_DENIED":
                        proposed = "QUARANTINED"
            except Exception as exc:
                parse_status = "PARSE_OR_FINGERPRINT_FAILED"
                reasons.append(f"FINGERPRINT_FAILED:{type(exc).__name__}")
                if proposed != "AUTO_DENIED":
                    proposed = "PARSER_REQUIRED"
    elif sheet_row and not worksheet_name:
        reasons.append("WORKSHEET_NAME_MISSING")
        if proposed != "AUTO_DENIED":
            proposed = "PARSER_REQUIRED"

    ids = detected_ids(file_row.get("relative_path", ""), sheet_row.get("stale_evidence", "") if sheet_row else "")
    project_id = file_row.get("project_id") or (ids[0] if ids else "")
    if sheet_row and sheet_row.get("detected_project_id") and project_id and sheet_row["detected_project_id"] != project_id:
        reasons.append("CROSS_PROJECT_IDENTITY")
        proposed = "AUTO_DENIED"

    if proposed == "NEEDS_HUMAN_REVIEW" and sheet_row and sheet_row.get("stale_template_status") == "CURRENT_PROJECT_ID_MATCH":
        proposed = "CANDIDATE"
        reasons.append("CURRENT_PROJECT_ID_MATCH_REVIEWABLE")
    if not reasons:
        reasons.append("REVIEW_REQUIRED_NO_AUTOMATIC_ALLOW")

    return {
        "decision_id": decision_id(file_row["file_id"], sheet_row.get("sheet_id") if sheet_row else None),
        "source_root_id": file_row.get("source_root", SOURCE_ROOT_ID),
        "project_id": project_id,
        "family_id": family_id(file_row.get("project_name", "")),
        "relative_path": file_row.get("relative_path", ""),
        "file_name": file_row.get("file_name", ""),
        "file_id": file_row.get("file_id", ""),
        "sheet_id": sheet_row.get("sheet_id", "") if sheet_row else "",
        "file_sha256": file_row.get("sha256", ""),
        "file_size": file_row.get("size_bytes", ""),
        "workbook_format": ext.lstrip("."),
        "worksheet_name": worksheet_name,
        "worksheet_visibility": worksheet_visibility,
        "worksheet_fingerprint": worksheet_fp,
        "workbook_sheet_count": sheet_count,
        "detected_project_identifiers": ids,
        "detected_customer_identifiers": [file_row.get("customer", "")] if file_row.get("customer") else [],
        "detected_document_role": file_row.get("primary_role", ""),
        "parser_name": parser_name,
        "parser_version": parser_version,
        "parse_status": parse_status,
        "formula_dependency_status": formula_status,
        "external_link_status": external_status,
        "named_range_status": named_range_status,
        "revision_status": revision_status,
        "duplicate_or_supersession": duplicate,
        "proposed_decision": proposed,
        "final_decision": "UNREVIEWED",
        "reason_codes": sorted(set(reasons)),
        "evidence_references": [file_row.get("file_id", ""), sheet_row.get("sheet_id", "") if sheet_row else ""],
        "policy_version": POLICY_VERSION,
        "reviewer_or_approver": "",
        "decision_timestamp": "",
        "approval_manifest_hash": "",
        "sanitized_artifact_hash": "",
    }


DECISION_FIELDS = [
    "decision_id", "source_root_id", "project_id", "family_id", "relative_path",
    "file_name", "file_id", "sheet_id", "file_sha256", "file_size",
    "workbook_format", "worksheet_name", "worksheet_visibility",
    "worksheet_fingerprint", "workbook_sheet_count",
    "detected_project_identifiers", "detected_customer_identifiers",
    "detected_document_role", "parser_name", "parser_version", "parse_status",
    "formula_dependency_status", "external_link_status", "named_range_status",
    "revision_status", "duplicate_or_supersession", "proposed_decision",
    "final_decision", "reason_codes", "evidence_references", "policy_version",
    "reviewer_or_approver", "decision_timestamp", "approval_manifest_hash",
    "sanitized_artifact_hash",
]


def flatten_decision(decision: dict[str, Any]) -> dict[str, Any]:
    row = dict(decision)
    for key in ["detected_project_identifiers", "detected_customer_identifiers", "reason_codes", "evidence_references"]:
        row[key] = "|".join(str(x) for x in row.get(key, []))
    return row


def load_inventory(file_index: Path, sheet_index: Path) -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    files = {row["file_id"]: row for row in csv_rows(file_index)}
    sheets = csv_rows(sheet_index)
    return files, sheets


def decisions_for_project(project_id: str, file_index: Path, sheet_index: Path, compute_fingerprints: bool = True) -> list[dict[str, Any]]:
    files, sheets = load_inventory(file_index, sheet_index)
    project_files = {fid: row for fid, row in files.items() if row.get("project_id") == project_id and row.get("extension") in {".xlsx", ".xlsm", ".xls"} and row.get("generator_eligibility") != "FORBIDDEN"}
    by_file: dict[str, list[dict[str, str]]] = {}
    for sheet in sheets:
        if sheet.get("workbook_file_id") in project_files:
            by_file.setdefault(sheet["workbook_file_id"], []).append(sheet)
    decisions: list[dict[str, Any]] = []
    for file_id, file_row in sorted(project_files.items()):
        file_sheets = by_file.get(file_id)
        if not file_sheets:
            decisions.append(decide_source_item(file_row, None, compute_fingerprints))
        else:
            for sheet in file_sheets:
                decisions.append(decide_source_item(file_row, sheet, compute_fingerprints))
    return decisions


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply source-guard decisions to inventory rows.")
    parser.add_argument("--project-id")
    parser.add_argument("--file-index", type=Path, default=Path("manifests/all_projects_file_role_index.csv"))
    parser.add_argument("--sheet-index", type=Path, default=Path("manifests/all_projects_workbook_sheet_index.csv"))
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--no-fingerprints", action="store_true")
    args = parser.parse_args()
    if not args.project_id:
        raise SystemExit("--project-id is required")
    decisions = decisions_for_project(args.project_id, args.file_index, args.sheet_index, not args.no_fingerprints)
    rows = [flatten_decision(d) for d in decisions]
    csv_write(args.output, rows, DECISION_FIELDS)
    write_json(args.output.with_suffix(".summary.json"), {
        "project_id": args.project_id,
        "decision_count": len(rows),
        "counts": {state: sum(1 for row in rows if row["proposed_decision"] == state) for state in sorted({row["proposed_decision"] for row in rows})},
        "policy_version": POLICY_VERSION,
    })
    print(f"wrote {len(rows)} decisions")


if __name__ == "__main__":
    main()
