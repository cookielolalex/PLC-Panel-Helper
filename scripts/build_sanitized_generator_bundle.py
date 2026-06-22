from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from harness_lib import sha256_file, sha256_json, write_json
from source_guard import SOURCE_ROOT_PATH, csv_rows, source_path_from_row, worksheet_fingerprint


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def sanitize_sheet_to_csv(workbook_path: Path, sheet_name: str, output_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows():
        values = []
        for cell in row:
            value = cell.value
            if value is None:
                values.append("")
            else:
                values.append(str(value).replace(str(SOURCE_ROOT_PATH), "[SRC-ALL-PROJECTS]"))
        if any(values):
            rows.append(values)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a sanitized generator bundle from validated approvals.")
    parser.add_argument("--decisions", type=Path, required=True)
    parser.add_argument("--approval-manifest", type=Path, required=True)
    parser.add_argument("--file-index", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--policy-version", default="source_guard_policy_v1")
    args = parser.parse_args()

    approval = __import__("json").loads(args.approval_manifest.read_text(encoding="utf-8"))
    if approval.get("status") != "PASS":
        raise SystemExit("approval manifest did not pass validation")
    decisions = {row["decision_id"]: row for row in csv_rows(args.decisions)}
    files = {row["file_id"]: row for row in csv_rows(args.file_index)}
    approved = [row for row in approval.get("decisions", []) if row.get("human_decision") == "HUMAN_APPROVED"]
    if not approved:
        raise SystemExit("no HUMAN_APPROVED source items")

    bundle_id = f"bundle-{args.project_id}"
    args.output_dir.mkdir(parents=True, exist_ok=True)
    sanitized_dir = args.output_dir / "sanitized_inputs"
    previews_dir = args.output_dir / "previews"
    artifacts = []
    provenance_rows = []
    for item in approved:
        decision = decisions.get(item["decision_id"])
        if not decision:
            raise SystemExit(f"missing decision {item['decision_id']}")
        if decision.get("project_id") != args.project_id:
            raise SystemExit(f"decision project mismatch {item['decision_id']}")
        file_row = files[decision["file_id"]]
        workbook_path = source_path_from_row(file_row)
        if sha256_file(workbook_path).upper() != decision["file_sha256"].upper():
            raise SystemExit(f"source hash changed {item['decision_id']}")
        fp = worksheet_fingerprint(workbook_path, decision["worksheet_name"])
        if fp["fingerprint"] != decision["worksheet_fingerprint"]:
            raise SystemExit(f"worksheet fingerprint changed {item['decision_id']}")
        safe_name = f"{decision['decision_id']}_{decision['worksheet_name']}".replace("/", "_").replace("\\", "_")
        artifact_path = sanitized_dir / f"{safe_name}.csv"
        row_count = sanitize_sheet_to_csv(workbook_path, decision["worksheet_name"], artifact_path)
        preview_path = previews_dir / f"{safe_name}.md"
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_text(
            f"# Sanitized Preview\n\nDecision: `{decision['decision_id']}`\nWorksheet: `{decision['worksheet_name']}`\nRows: {row_count}\n",
            encoding="utf-8",
            newline="\n",
        )
        artifacts.append({
            "artifact_id": decision["decision_id"],
            "path": str(artifact_path.relative_to(args.output_dir)),
            "sha256": sha256_file(artifact_path),
            "source_decision_id": decision["decision_id"],
        })
        provenance_rows.append({
            "source_decision_id": decision["decision_id"],
            "source_file_id": decision["file_id"],
            "source_sheet_id": decision["sheet_id"],
            "sanitized_artifact": str(artifact_path.relative_to(args.output_dir)),
        })

    bundle_manifest = {
        "bundle_id": bundle_id,
        "project_id": args.project_id,
        "policy_version": args.policy_version,
        "status": "BUILT_PENDING_VERIFICATION",
        "artifacts": artifacts,
    }
    visible_file_manifest = {
        "bundle_id": bundle_id,
        "files": [{"path": artifact["path"], "sha256": artifact["sha256"]} for artifact in artifacts],
    }
    write_json(args.output_dir / "bundle_manifest.json", bundle_manifest)
    write_json(args.output_dir / "approval_manifest.json", approval)
    write_json(args.output_dir / "visible_file_manifest.json", visible_file_manifest)
    write_json(args.output_dir / "provenance_map.json", {"bundle_id": bundle_id, "rows": provenance_rows})
    write_json(args.output_dir / "verification_results.json", {
        "status": "PENDING_VERIFICATION",
        "errors": ["BUNDLE_NOT_VERIFIED"],
        "warnings": [],
        "artifact_count": len(artifacts),
    })
    write_json(args.output_dir / "bundle_hashes.json", {
        "bundle_manifest": sha256_file(args.output_dir / "bundle_manifest.json"),
        "approval_manifest": sha256_file(args.output_dir / "approval_manifest.json"),
        "visible_file_manifest": sha256_file(args.output_dir / "visible_file_manifest.json"),
        "provenance_map": sha256_file(args.output_dir / "provenance_map.json"),
        "verification_results": sha256_file(args.output_dir / "verification_results.json"),
        "artifacts": artifacts,
    })
    print(f"built {bundle_id}")


if __name__ == "__main__":
    main()
