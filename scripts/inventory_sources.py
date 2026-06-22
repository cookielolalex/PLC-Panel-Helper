from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

from harness_lib import classify_path, detect_forbidden_text, sha256_file, write_json


WORKBOOK_EXTS = {".xlsx", ".xlsm", ".xls"}
PROJECT_ID_RE = re.compile(r"(?<!\d)(1[0-9]{2}[0-9]{4})(?!\d)")
DATE_RE = re.compile(r"(1[0-9]{2}[./-]\d{1,2}[./-]\d{1,2}|\d{4}[./-]\d{1,2}[./-]\d{1,2})")


FILE_FIELDS = [
    "file_id",
    "source_root",
    "relative_path",
    "archive_member_path",
    "project_id",
    "project_name",
    "customer",
    "file_name",
    "extension",
    "size_bytes",
    "modified_time",
    "sha256",
    "content_fingerprint",
    "duplicate_of",
    "primary_role",
    "secondary_tags",
    "drawing_output_type",
    "generator_eligibility",
    "forbidden_signature_match",
    "classification_confidence",
    "basis",
    "parse_status",
    "review_status",
    "notes",
    "absolute_path",
]

SHEET_FIELDS = [
    "sheet_id",
    "workbook_file_id",
    "sheet_name",
    "visibility",
    "used_range",
    "print_area",
    "named_ranges",
    "external_links",
    "detected_project_id",
    "detected_customer",
    "detected_project_name",
    "detected_date",
    "detected_revision",
    "stale_template_status",
    "stale_evidence",
    "generator_eligibility",
    "approved_by",
    "approval_time",
    "parse_status",
    "notes",
]

PROJECT_FIELDS = [
    "project_id",
    "project_path",
    "project_name",
    "file_count",
    "workbook_count",
    "allowed_contract_workbook_count",
    "allowed_material_list_workbook_count",
    "allowed_supporting_workbook_count",
    "forbidden_production_control_file_count",
    "forbidden_electrical_drawing_count",
    "completed_production_drawing_reference_count",
    "completed_sheetmetal_drawing_reference_count",
    "completed_punch_drawing_reference_count",
    "has_all_three_completed_targets",
    "has_forbidden_sources",
    "generator_candidate_status",
    "notes",
]


def csv_write(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def extract_project_from_rel(relative: Path) -> tuple[str | None, str | None, str | None]:
    parts = relative.parts
    for part in parts:
        match = PROJECT_ID_RE.search(part)
        if match:
            project_id = match.group(1)
            project_name = part
            customer = None
            if "-" in part:
                customer = part.split("-", 1)[0].replace(project_id, "").strip() or None
            return project_id, project_name, customer
    return None, None, None


def stable_file_id(source_root_id: str, relative: str) -> str:
    import hashlib

    suffix = hashlib.sha1(relative.encode("utf-8")).hexdigest()[:12].upper()
    return f"{source_root_id}-FILE-{suffix}"


def row_for_file(root: Path, path: Path, source_root_id: str, hash_files: bool) -> dict[str, Any]:
    rel = path.relative_to(root)
    rel_text = str(rel)
    project_id, project_name, customer = extract_project_from_rel(rel)
    classified = classify_path(rel_text)
    digest = sha256_file(path) if hash_files else "NOT_HASHED"
    return {
        "file_id": stable_file_id(source_root_id, rel_text),
        "source_root": source_root_id,
        "relative_path": rel_text,
        "archive_member_path": "",
        "project_id": project_id or "",
        "project_name": project_name or "",
        "customer": customer or "",
        "file_name": path.name,
        "extension": path.suffix.lower(),
        "size_bytes": path.stat().st_size,
        "modified_time": path.stat().st_mtime,
        "sha256": digest,
        "content_fingerprint": digest if hash_files else "",
        "duplicate_of": "",
        "primary_role": classified["primary_role"],
        "secondary_tags": "|".join(classified["forbidden_hits"]),
        "drawing_output_type": classified["drawing_output_type"] or "",
        "generator_eligibility": classified["generator_eligibility"],
        "forbidden_signature_match": classified["forbidden_signature_match"],
        "classification_confidence": classified["classification_confidence"],
        "basis": classified["basis"],
        "parse_status": "NOT_PARSED",
        "review_status": "BOOTSTRAP_CLASSIFIED",
        "notes": "",
        "absolute_path": str(path),
    }


def workbook_text_sample(ws: Any, max_rows: int = 40, max_cols: int = 20) -> str:
    values: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_rows), max_col=min(ws.max_column or 1, max_cols)):
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return " ".join(values)


def detect_sheet_identity(text: str) -> dict[str, str]:
    projects = PROJECT_ID_RE.findall(text)
    dates = DATE_RE.findall(text)
    return {
        "project_id": projects[0] if projects else "",
        "date": dates[0] if dates else "",
        "all_project_ids": "|".join(sorted(set(projects))),
    }


def inspect_workbook(path: Path, file_row: dict[str, Any]) -> list[dict[str, Any]]:
    ext = path.suffix.lower()
    if ext == ".xls":
        return [{
            "sheet_id": f"{file_row['file_id']}-XLS-UNPARSED",
            "workbook_file_id": file_row["file_id"],
            "sheet_name": "",
            "visibility": "unknown",
            "used_range": "",
            "print_area": "",
            "named_ranges": "",
            "external_links": "",
            "detected_project_id": "",
            "detected_customer": "",
            "detected_project_name": "",
            "detected_date": "",
            "detected_revision": "",
            "stale_template_status": "HUMAN_REVIEW_REQUIRED",
            "stale_evidence": "Legacy .xls parser not available in Phase 1 bootstrap scanner.",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "approved_by": "",
            "approval_time": "",
            "parse_status": "UNSUPPORTED_XLS_PARSER_REQUIRED",
            "notes": "",
        }]

    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    except Exception as exc:
        return [{
            "sheet_id": f"{file_row['file_id']}-PARSE-FAILED",
            "workbook_file_id": file_row["file_id"],
            "sheet_name": "",
            "visibility": "unknown",
            "used_range": "",
            "print_area": "",
            "named_ranges": "",
            "external_links": "",
            "detected_project_id": "",
            "detected_customer": "",
            "detected_project_name": "",
            "detected_date": "",
            "detected_revision": "",
            "stale_template_status": "HUMAN_REVIEW_REQUIRED",
            "stale_evidence": f"Workbook parse failed: {type(exc).__name__}: {exc}",
            "generator_eligibility": "HUMAN_REVIEW_REQUIRED",
            "approved_by": "",
            "approval_time": "",
            "parse_status": "PARSE_FAILED",
            "notes": "",
        }]

    external_links = "|".join(str(x) for x in getattr(wb, "_external_links", []))
    rows: list[dict[str, Any]] = []
    workbook_project = file_row.get("project_id", "")
    forbidden_file = file_row.get("generator_eligibility") == "FORBIDDEN"
    for idx, ws in enumerate(wb.worksheets, start=1):
        sample = workbook_text_sample(ws)
        identity = detect_sheet_identity(sample)
        hits = detect_forbidden_text(ws.title + " " + sample[:2000])
        stale = "INSUFFICIENT_IDENTITY"
        eligibility = "HUMAN_REVIEW_REQUIRED"
        evidence = ""
        detected_project = identity["project_id"]
        if forbidden_file or hits:
            stale = "FORBIDDEN_SOURCE"
            eligibility = "FORBIDDEN"
            evidence = "Forbidden label in file or worksheet content."
        elif detected_project and workbook_project and detected_project != workbook_project:
            stale = "STALE_TEMPLATE_SHEET"
            eligibility = "FORBIDDEN"
            evidence = f"Worksheet project {detected_project} differs from workbook path project {workbook_project}."
        elif detected_project and (not workbook_project or detected_project == workbook_project):
            stale = "CURRENT_PROJECT_ID_MATCH"
            eligibility = "HUMAN_REVIEW_REQUIRED"
            evidence = "Current project identity detected; positive approval still required."
        elif ws.sheet_state != "visible":
            stale = "HIDDEN_SHEET_REVIEW_REQUIRED"
            evidence = "Hidden or very hidden sheet lacks sufficient current identity."

        rows.append({
            "sheet_id": f"{file_row['file_id']}-SHEET-{idx:03d}",
            "workbook_file_id": file_row["file_id"],
            "sheet_name": ws.title,
            "visibility": ws.sheet_state,
            "used_range": ws.calculate_dimension(),
            "print_area": str(ws.print_area) if ws.print_area else "",
            "named_ranges": "",
            "external_links": external_links,
            "detected_project_id": detected_project,
            "detected_customer": "",
            "detected_project_name": "",
            "detected_date": identity["date"],
            "detected_revision": "",
            "stale_template_status": stale,
            "stale_evidence": evidence or sample[:300].replace("\n", " "),
            "generator_eligibility": eligibility,
            "approved_by": "",
            "approval_time": "",
            "parse_status": "PARSED",
            "notes": f"all_project_ids={identity['all_project_ids']}",
        })
    return rows


def project_rows(file_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_project: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in file_rows:
        project_id = row.get("project_id") or "UNASSIGNED"
        by_project[project_id].append(row)

    out: list[dict[str, Any]] = []
    role_fields = {
        "allowed_contract_workbook": "allowed_contract_workbook_count",
        "allowed_material_list_workbook": "allowed_material_list_workbook_count",
        "allowed_supporting_workbook": "allowed_supporting_workbook_count",
        "forbidden_production_control_file": "forbidden_production_control_file_count",
        "forbidden_electrical_drawing": "forbidden_electrical_drawing_count",
        "completed_production_drawing_reference": "completed_production_drawing_reference_count",
        "completed_sheetmetal_drawing_reference": "completed_sheetmetal_drawing_reference_count",
        "completed_punch_drawing_reference": "completed_punch_drawing_reference_count",
    }
    for project_id, rows in sorted(by_project.items()):
        counts = {field: 0 for field in role_fields.values()}
        for row in rows:
            field = role_fields.get(row.get("primary_role"))
            if field:
                counts[field] += 1
        target_complete = all(counts[k] > 0 for k in [
            "completed_production_drawing_reference_count",
            "completed_sheetmetal_drawing_reference_count",
            "completed_punch_drawing_reference_count",
        ])
        has_forbidden = counts["forbidden_production_control_file_count"] > 0 or counts["forbidden_electrical_drawing_count"] > 0
        workbook_count = sum(1 for r in rows if r.get("extension") in WORKBOOK_EXTS)
        candidate = "NO_APPROVED_WORKSHEETS"
        if workbook_count and target_complete:
            candidate = "REFERENCE_COMPLETE_WORKSHEET_REVIEW_REQUIRED"
        elif workbook_count:
            candidate = "WORKSHEET_REVIEW_REQUIRED"
        sample = rows[0]
        out.append({
            "project_id": project_id,
            "project_path": str(Path(sample["relative_path"]).parent) if project_id != "UNASSIGNED" else "",
            "project_name": sample.get("project_name", ""),
            "file_count": len(rows),
            "workbook_count": workbook_count,
            **counts,
            "has_all_three_completed_targets": target_complete,
            "has_forbidden_sources": has_forbidden,
            "generator_candidate_status": candidate,
            "notes": "",
        })
    return out


def duplicate_mark(file_rows: list[dict[str, Any]]) -> None:
    first_by_hash: dict[str, str] = {}
    for row in file_rows:
        digest = row.get("sha256")
        if not digest or digest == "NOT_HASHED":
            continue
        if digest in first_by_hash:
            row["duplicate_of"] = first_by_hash[digest]
        else:
            first_by_hash[digest] = row["file_id"]


def filter_active_year(file_rows: list[dict[str, Any]], active_year: str) -> list[dict[str, Any]]:
    prefix = f"{active_year}年度工作"
    return [row for row in file_rows if row["relative_path"].startswith(prefix)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Create read-only file, workbook-sheet, and project manifests.")
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--source-root-id", default="SRC-ALL-PROJECTS")
    parser.add_argument("--output-prefix", default="all_projects")
    parser.add_argument("--active-year", default="115")
    parser.add_argument("--hash-files", action="store_true")
    parser.add_argument("--inspect-workbooks", action="store_true")
    parser.add_argument("--max-workbooks", type=int, default=0, help="0 means inspect all workbooks.")
    parser.add_argument("--progress-every", type=int, default=250)
    args = parser.parse_args()

    start = time.time()
    if not args.root.exists() or not args.root.is_dir():
        raise SystemExit(f"source root is missing or not a directory: {args.root}")

    paths = sorted(p for p in args.root.rglob("*") if p.is_file())
    file_rows: list[dict[str, Any]] = []
    for index, path in enumerate(paths, start=1):
        file_rows.append(row_for_file(args.root, path, args.source_root_id, args.hash_files))
        if args.progress_every and index % args.progress_every == 0:
            print(f"indexed {index}/{len(paths)} files", file=sys.stderr)
    duplicate_mark(file_rows)

    manifest_dir = Path("manifests")
    reports_dir = Path("reports")
    all_file_csv = manifest_dir / f"{args.output_prefix}_file_role_index.csv"
    csv_write(all_file_csv, file_rows, FILE_FIELDS)

    active_rows = filter_active_year(file_rows, args.active_year)
    csv_write(manifest_dir / f"{args.active_year}_file_role_index.csv", active_rows, FILE_FIELDS)

    workbook_rows: list[dict[str, Any]] = []
    if args.inspect_workbooks:
        workbook_file_rows = [row for row in file_rows if row["extension"] in WORKBOOK_EXTS]
        if args.max_workbooks:
            workbook_file_rows = workbook_file_rows[: args.max_workbooks]
        for index, file_row in enumerate(workbook_file_rows, start=1):
            path = Path(file_row["absolute_path"])
            workbook_rows.extend(inspect_workbook(path, file_row))
            if args.progress_every and index % max(1, args.progress_every // 10) == 0:
                print(f"inspected {index}/{len(workbook_file_rows)} workbooks", file=sys.stderr)
    csv_write(manifest_dir / f"{args.output_prefix}_workbook_sheet_index.csv", workbook_rows, SHEET_FIELDS)
    active_file_ids = {row["file_id"] for row in active_rows}
    active_sheet_rows = [row for row in workbook_rows if row["workbook_file_id"] in active_file_ids]
    csv_write(manifest_dir / f"{args.active_year}_workbook_sheet_index.csv", active_sheet_rows, SHEET_FIELDS)

    projects = project_rows(file_rows)
    csv_write(manifest_dir / f"{args.output_prefix}_project_manifest.csv", projects, PROJECT_FIELDS)
    csv_write(manifest_dir / f"{args.active_year}_project_manifest.csv", [p for p in projects if str(p["project_id"]).startswith(args.active_year)], PROJECT_FIELDS)

    stale_rows = [row for row in workbook_rows if row["stale_template_status"] in {"STALE_TEMPLATE_SHEET", "HIDDEN_SHEET_REVIEW_REQUIRED", "INSUFFICIENT_IDENTITY"}]
    csv_write(reports_dir / "stale_sheet_review_queue.csv", stale_rows, SHEET_FIELDS)
    manual_rows = []
    for row in workbook_rows:
        if row["generator_eligibility"] == "HUMAN_REVIEW_REQUIRED":
            manual_rows.append({
                "queue_id": f"WS-{len(manual_rows)+1:06d}",
                "item_type": "worksheet",
                "item_id": row["sheet_id"],
                "reason": row["stale_template_status"],
                "status": "HUMAN_REVIEW_REQUIRED",
            })
    csv_write(reports_dir / "manual_review_queue.csv", manual_rows, ["queue_id", "item_type", "item_id", "reason", "status"])

    summary = {
        "root": str(args.root),
        "source_root_id": args.source_root_id,
        "file_count": len(file_rows),
        "active_year_file_count": len(active_rows),
        "workbook_sheet_rows": len(workbook_rows),
        "active_year_sheet_rows": len(active_sheet_rows),
        "project_count": len([p for p in projects if p["project_id"] != "UNASSIGNED"]),
        "hash_files": args.hash_files,
        "inspect_workbooks": args.inspect_workbooks,
        "duration_seconds": round(time.time() - start, 2),
        "outputs": [
            str(all_file_csv),
            str(manifest_dir / f"{args.active_year}_file_role_index.csv"),
            str(manifest_dir / f"{args.output_prefix}_workbook_sheet_index.csv"),
            str(manifest_dir / f"{args.active_year}_workbook_sheet_index.csv"),
            str(manifest_dir / f"{args.output_prefix}_project_manifest.csv"),
            str(manifest_dir / f"{args.active_year}_project_manifest.csv"),
        ],
    }
    write_json(manifest_dir / f"{args.output_prefix}_inventory_summary.json", summary)
    print(summary)


if __name__ == "__main__":
    main()
